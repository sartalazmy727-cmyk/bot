# bot.py
# -*- coding: utf-8 -*-

"""
AOU Kuwait Telegram Bot â€” FINAL (Updated)

âœ… Ø¨ÙˆØ§Ø¨Ø© Ø¬Ù‡Ø© Ø§ØªØµØ§Ù„ (Ù„Ù„Ø£Ø¹Ø¶Ø§Ø¡ ÙÙ‚Ø·) + ÙŠÙ‚Ø¨Ù„ Ø§Ù„ÙƒÙˆÙŠØª ÙÙ‚Ø· (+965)
âœ… Ø§Ù„Ø³ÙˆØ¨Ø± Ø£Ø¯Ù…Ù† ÙÙ‚Ø· (IDs Ù…Ø­Ø¯Ø¯Ø©) ÙŠÙ‚Ø¯Ø±:
   - ØªØ´ØºÙŠÙ„/Ø¥ÙŠÙ‚Ø§Ù Ø®Ø¯Ù…Ø© Ø§Ù„Ø£Ø±Ù‚Ø§Ù…
   - Ù…Ø³Ø­ Ù‚ÙˆØ§Ø¦Ù… Ø§Ù„Ø£Ø±Ù‚Ø§Ù… (Ø§Ù„Ù…Ù‚Ø¨ÙˆÙ„Ø©/Ø§Ù„Ù…Ø±ÙÙˆØ¶Ø©)
   - Ø¥Ø®ÙØ§Ø¡/Ø¥Ø¸Ù‡Ø§Ø± Ø§Ù„Ø£Ø²Ø±Ø§Ø± Ù„Ù„Ø¹Ø§Ù…Ø©
   - ØªØºÙŠÙŠØ± Ø£Ø³Ù…Ø§Ø¡ Ø§Ù„Ø£Ø²Ø±Ø§Ø±
   - Ø¥Ø¯Ø§Ø±Ø© Ø§Ù„Ø£Ø¯Ù…Ù† Ø§Ù„Ø¥Ø¶Ø§ÙÙŠÙŠÙ†
   - Ø§Ù„ØªØ­ÙƒÙ… Ø§Ù„ÙƒØ§Ù…Ù„ Ø¨Ø§Ù„ØªÙ‚ÙˆÙŠÙ… (ØªÙ„Ù‚Ø§Ø¦ÙŠ/Ù…Ø®ØµØµ/Ø­Ø°Ù) + (Ø¥Ø±ÙØ§Ù‚ PDF/ØµÙˆØ±Ø©/Ù†Øµ)
âœ… Ø§Ù„Ø£Ø¯Ù…Ù† Ø§Ù„Ø¥Ø¶Ø§ÙÙŠ ÙŠÙ‚Ø¯Ø±:
   - Ø±ÙØ¹/ØªØ¹Ø¯ÙŠÙ„/Ø­Ø°Ù (Ù…Ù„Ù/ØµÙˆØ±Ø©/Ù†Øµ) Ù„Ø£Ù‚Ø³Ø§Ù… Ø§Ù„Ø®Ø¯Ù…Ø§Øª ÙˆØ§Ù„Ù…Ù„Ø®ØµØ§Øª
   - Ø¥Ø¶Ø§ÙØ©/ØªØ¹Ø¯ÙŠÙ„/Ø­Ø°Ù Ø§Ù„Ù‚Ø±ÙˆØ¨Ø§Øª Ø§Ù„Ø¹Ø§Ù…Ø©
   - Ø¥Ø¶Ø§ÙØ©/Ø­Ø°Ù Ø±ÙˆØ§Ø¨Ø· Ù‚Ø±ÙˆØ¨Ø§Øª Ø§Ù„ÙƒÙ„ÙŠØ§Øª
   - Ù…Ø´Ø§Ù‡Ø¯Ø© Ø§Ù„Ø£Ø±Ù‚Ø§Ù… Ø§Ù„Ù…Ù‚Ø¨ÙˆÙ„Ø©/Ø§Ù„Ù…Ø±ÙÙˆØ¶Ø© + ØªØµØ¯ÙŠØ± Excel (Ø¨Ø¯ÙˆÙ† Ù…Ø³Ø­/ØªØºÙŠÙŠØ±)
âœ… Ø²Ø± "Ø§Ù„ØªÙ‚ÙˆÙŠÙ… Ø§Ù„Ø¬Ø§Ù…Ø¹ÙŠ": ÙŠØ³Ù…Ø­ Ù„Ù„Ø£Ø¯Ù…Ù† Ø¨Ø¥Ø±ÙØ§Ù‚ (PDF/ØµÙˆØ±Ø©/Ù†Øµ) Ø¯Ø§Ø®Ù„ Ø§Ù„Ø²Ø± Ù†ÙØ³Ù‡.
âœ… Ø²Ø± "Ø§Ù„Ø³Ø­Ø¨ ÙˆØ§Ù„Ø¥Ø¶Ø§ÙØ©": Ø£ØµØ¨Ø­ Ù‚Ø§Ø¨Ù„Ø§Ù‹ Ù„Ù„Ø¥Ø¯Ø§Ø±Ø© Ø¨ÙˆØ§Ø³Ø·Ø© Ø§Ù„Ø£Ø¯Ù…Ù† ÙÙ‚Ø· (PDF/ØµÙˆØ±Ø©/Ù†Øµ).
âœ… ØªØµØ¯ÙŠØ± Ø§Ù„Ø£Ø±Ù‚Ø§Ù… Ø¥Ù„Ù‰ Ù…Ù„Ù Excel (Sheet Ù„Ù„Ù…Ù‚Ø¨ÙˆÙ„Ø© ÙˆSheet Ù„Ù„Ù…Ø±ÙÙˆØ¶Ø©)

Ø§Ù„Ù…ØªØ·Ù„Ø¨Ø§Øª:
  pip install python-telegram-bot==21.6 openpyxl

Ù…Ù„Ø§Ø­Ø¸Ø©:
- Ø¶Ø¹ ØªÙˆÙƒÙ†Ùƒ Ø¨Ø¯Ù„ "123456789"
- SUPER_ADMIN_IDS ØªÙ… ØªØ«Ø¨ÙŠØªÙ‡Ø§ Ø­Ø³Ø¨ Ø·Ù„Ø¨Ùƒ
"""

from __future__ import annotations

import asyncio
import json
import logging
import os
import re
import time
import urllib.request
from dataclasses import dataclass
from threading import RLock
from typing import Any, Dict, List, Optional, Set, Tuple

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from telegram import (
    BotCommand,
    BotCommandScopeChat,
    BotCommandScopeDefault,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    KeyboardButton,
    ReplyKeyboardMarkup,
    Update,
)
from telegram.constants import ChatType
from telegram.error import BadRequest
from telegram.ext import (
    Application,
    CommandHandler,
    ContextTypes,
    MessageHandler,
    filters,
)

# =========================================================
# BASIC SETTINGS
# =========================================================
TOKEN = "8308362115:AAFj9WDYSjF0YYlvo1r1bgkRPyXi49h1VJ4"  # <- ØºÙŠÙ‘Ø±Ù‡
DATA_FILE = "bot_data.json"
DATA_LOCK = RLock()

logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO,
)
LOG = logging.getLogger("AOU_BOT_FINAL")

MAX_TEXT_LEN = 3600
MAX_INBOX = 300

# âœ… Ø§Ù„Ø³ÙˆØ¨Ø± Ø£Ø¯Ù…Ù† ÙÙ‚Ø· (Ø­Ø³Ø¨ Ø·Ù„Ø¨Ùƒ)
SUPER_ADMIN_IDS: Set[int] = {8136678328, 164796308, 8318266324}

# =========================================================
# LINKS
# =========================================================
UNIV_URL = "https://www.aou.edu.kw/ar/Pages/default.aspx"
ADMISSION_URL = "https://www.aou.edu.kw/ar/admission/pages/undergraduate-apply.aspx"
AOU_CALENDAR_PAGE_EN = "https://www.aou.edu.kw/students/Pages/academic-calendar.aspx"
AOU_CALENDAR_PAGE_AR = "https://www.aou.edu.kw/ar/students/Pages/academic-calendar.aspx"
AOU_SOCIAL_LINKTREE = "https://linktr.ee/aou_kw"

# =========================================================
# TEXTS
# =========================================================
WELCOME_TEXT = (
    "Ù…Ø±Ø­Ø¨Ù‹Ø§ Ø¨Ùƒ ğŸ‘‹\n"
    "Ø£Ù†Ø§ Ø¨ÙˆØª Ø§Ù„Ø¬Ø§Ù…Ø¹Ø© Ø§Ù„Ø¹Ø±Ø¨ÙŠØ© Ø§Ù„Ù…ÙØªÙˆØ­Ø© (AOU) â€“ Ø§Ù„ÙƒÙˆÙŠØª ğŸ‡°ğŸ‡¼\n"
    "Ø£Ø³Ø§Ø¹Ø¯Ùƒ ÙÙŠ Ø§Ù„Ø§Ø³ØªÙØ³Ø§Ø±Ø§Øª Ø§Ù„Ø£ÙƒØ§Ø¯ÙŠÙ…ÙŠØ©ØŒ Ø§Ù„Ù‚Ø¨ÙˆÙ„ØŒ Ø§Ù„Ø¬Ø¯Ø§ÙˆÙ„ØŒ ÙˆØ§Ù„Ø®Ø¯Ù…Ø§Øª Ø§Ù„Ø·Ù„Ø§Ø¨ÙŠØ© ğŸ“š\n"
    "Ø§Ø®ØªØ± Ù…Ù† Ø§Ù„Ø£Ø²Ø±Ø§Ø± Ø¨Ø§Ù„Ø£Ø³ÙÙ„ âœ…"
)

CONTACT_PROMPT_TEXT = (
    "ğŸ“² Ù‚Ø¨Ù„ Ø§Ø³ØªØ®Ø¯Ø§Ù… Ø§Ù„Ø¨ÙˆØªØŒ Ø´Ø§Ø±Ùƒ Ø±Ù‚Ù…Ùƒ (Ø¬Ù‡Ø© Ø§Ù„Ø§ØªØµØ§Ù„).\n"
    "âœ… ÙŠÙ‚Ø¨Ù„ Ø£Ø±Ù‚Ø§Ù… Ø§Ù„ÙƒÙˆÙŠØª ÙÙ‚Ø· (+965).\n"
)

IMPORTANT_NOTICE_TEXT = (
    "ØªÙ†Ø¨ÙŠÙ‡ Ù…Ù‡Ù… ğŸ“Œ\n"
    "Ù‡Ø°Ø§ Ø§Ù„Ø¨ÙˆØª Ù…Ø®ØµØµ Ù„Ø·Ù„Ø¨Ø© ÙˆØ·Ø§Ù„Ø¨Ø§Øª Ø§Ù„ÙƒÙˆÙŠØª ÙÙ‚Ø· ğŸ‡°ğŸ‡¼\n"
    "ÙˆÙ„Ø¶Ù…Ø§Ù† Ù‚Ø¨ÙˆÙ„Ùƒ ÙÙŠ Ø§Ù„Ø¨ÙˆØª âœ…\n"
    "ÙŠØ±Ø¬Ù‰ Ù…Ø´Ø§Ø±ÙƒØ© Ø¬Ù‡Ø§Øª Ø§Ù„Ø§ØªØµØ§Ù„ Ø­ØªÙ‰ Ù†ØªÙ…ÙƒÙ† Ù…Ù† Ø§Ù„ØªØ­Ù‚Ù‚ Ø¨Ø³Ø±Ø¹Ø© ÙˆØªÙˆÙÙŠØ± Ø§Ù„Ø¬Ù‡Ø¯ ÙˆØ§Ù„ÙˆÙ‚Øª.\n"
    "Ø´ÙƒØ±Ù‹Ø§ Ù„ØªØ¹Ø§ÙˆÙ†ÙƒÙ… ğŸ¤"
)

AOU_ABOUT = (
    "ğŸ“ Ø§Ù„Ø¬Ø§Ù…Ø¹Ø© Ø§Ù„Ø¹Ø±Ø¨ÙŠØ© Ø§Ù„Ù…ÙØªÙˆØ­Ø© (AOU) â€“ ÙØ±Ø¹ Ø§Ù„ÙƒÙˆÙŠØª\n\n"
    "â€¢ ØªØ¹Ù„ÙŠÙ… Ù…Ø¯Ù…Ø¬ (Ø­Ø¶ÙˆØ±ÙŠ + Ø¥Ù„ÙƒØªØ±ÙˆÙ†ÙŠ)\n"
    "â€¢ Ù…Ø±ÙˆÙ†Ø© Ù…Ù†Ø§Ø³Ø¨Ø© Ù„Ù„Ø·Ù„Ø§Ø¨ ÙˆØ§Ù„Ù…ÙˆØ¸ÙÙŠÙ†\n\n"
    f"ğŸ”— Ù…ÙˆÙ‚Ø¹ Ø§Ù„Ø¬Ø§Ù…Ø¹Ø©: {UNIV_URL}"
)

AOU_ADMISSION = (
    "âœ… Ø§Ù„Ù‚Ø¨ÙˆÙ„ (Ù…Ø®ØªØµØ±)\n\n"
    "1) Ø´Ù‡Ø§Ø¯Ø© Ø§Ù„Ø«Ø§Ù†ÙˆÙŠØ© Ø§Ù„Ø¹Ø§Ù…Ø© Ø£Ùˆ Ù…Ø§ ÙŠØ¹Ø§Ø¯Ù„Ù‡Ø§.\n"
    "2) Ù†Ø³Ø¨Ø© Ù„Ø§ ØªÙ‚Ù„ Ø¹Ù† 60% Ø£Ùˆ GPA Ù„Ø§ ÙŠÙ‚Ù„ Ø¹Ù† 2.00.\n"
    "3) Ø¥Ø°Ø§ ÙƒØ§Ù† Ø§Ù„Ù…Ø¹Ø¯Ù„ Ø£Ù‚Ù„: ÙŠÙ…ÙƒÙ† Ø§Ù„ØªÙ‚Ø¯ÙŠÙ… Ø¨Ø´Ø±Ø· Ø®Ø¨Ø±Ø© Ø¹Ù…Ù„ 4 Ø³Ù†ÙˆØ§Øª Ø¨Ø¹Ø¯ Ø§Ù„Ø«Ø§Ù†ÙˆÙŠØ©.\n"
    "4) Ù„ØªØ®ØµØµ ØªÙ‚Ù†ÙŠØ© Ø§Ù„Ù…Ø¹Ù„ÙˆÙ…Ø§Øª (IT): ÙŠØ´ØªØ±Ø· Ø£Ù† ØªÙƒÙˆÙ† Ø§Ù„Ø«Ø§Ù†ÙˆÙŠØ© Ø¹Ù„Ù…ÙŠØ©.\n\n"
    f"ğŸ”— Ø±Ø§Ø¨Ø· Ø§Ù„Ù‚Ø¨ÙˆÙ„/Ø§Ù„ØªÙ‚Ø¯ÙŠÙ…: {ADMISSION_URL}"
)

AOU_MAJORS = (
    "ğŸ“š Ø§Ù„ØªØ®ØµØµØ§Øª (Ø¨ÙƒØ§Ù„ÙˆØ±ÙŠÙˆØ³) ÙÙŠ AOU â€“ Ø§Ù„ÙƒÙˆÙŠØª\n\n"
    "ğŸ¢ ÙƒÙ„ÙŠØ© Ø¥Ø¯Ø§Ø±Ø© Ø§Ù„Ø£Ø¹Ù…Ø§Ù„:\n"
    "â€¢ Ø§Ù„Ù…Ø­Ø§Ø³Ø¨Ø© â€¢ Ø§Ù„Ø§Ù‚ØªØµØ§Ø¯ â€¢ MIS â€¢ Ø§Ù„Ø¥Ø¯Ø§Ø±Ø© â€¢ Ø§Ù„ØªØ³ÙˆÙŠÙ‚ â€¢ Ø§Ù„Ù…Ù…Ø§Ø±Ø³Ø§Øª Ø§Ù„Ù†Ø¸Ù…ÙŠØ©\n\n"
    "ğŸ’» ÙƒÙ„ÙŠØ© Ø¯Ø±Ø§Ø³Ø§Øª Ø§Ù„Ø­Ø§Ø³ÙˆØ¨:\n"
    "â€¢ ITC\n\n"
    "ğŸ‘©â€ğŸ« ÙƒÙ„ÙŠØ© Ø§Ù„ØªØ±Ø¨ÙŠØ©:\n"
    "â€¢ Ø§Ù„ØªØ±Ø¨ÙŠØ© Ø§Ù„Ø§Ø¨ØªØ¯Ø§Ø¦ÙŠØ© â€¢ Ø§Ù„ØªØ±Ø¨ÙŠØ© Ø§Ù„Ø®Ø§ØµØ©\n\n"
    "ğŸŒ ÙƒÙ„ÙŠØ© Ø§Ù„Ù„ØºØ§Øª:\n"
    "â€¢ Ø¨Ø±Ø§Ù…Ø¬ Ù„ØºØ§Øª Ø­Ø³Ø¨ Ø§Ù„Ù…ØªØ§Ø­ Ø¨Ø§Ù„ÙØ±Ø¹."
)

# âœ… Ø£ØµØ¨Ø­ Ø²Ø± Ø§Ù„Ø³Ø­Ø¨ ÙˆØ§Ù„Ø¥Ø¶Ø§ÙØ© Ù‚Ø§Ø¨Ù„ Ù„Ù„Ø¥Ø¯Ø§Ø±Ø© (PDF/ØµÙˆØ±Ø©/Ù†Øµ) Ù„Ù„Ø£Ø¯Ù…Ù† ÙÙ‚Ø·
AOU_ADD_DROP_FALLBACK = (
    "âœï¸ Ø§Ù„Ø³Ø­Ø¨ ÙˆØ§Ù„Ø¥Ø¶Ø§ÙØ©\n\n"
    "âœ… Ø§Ù„Ø¥Ø¶Ø§ÙØ©: Ø®Ù„Ø§Ù„ ÙØªØ±Ø© Ø§Ù„Ø³Ø­Ø¨ ÙˆØ§Ù„Ø¥Ø¶Ø§ÙØ© Ø­Ø³Ø¨ Ø§Ù„ØªÙ‚ÙˆÙŠÙ….\n"
    "âœ… Ø§Ù„Ø³Ø­Ø¨: ÙŠØ­Ù‚ Ù„Ùƒ Ø§Ù„Ø§Ù†Ø³Ø­Ø§Ø¨ Ø®Ù„Ø§Ù„ Ø£ÙˆÙ„ 6 Ø£Ø³Ø§Ø¨ÙŠØ¹ Ù…Ù† Ø§Ù„ÙØµÙ„.\n\n"
    "ğŸ’° Ø§Ù„Ø§Ø³ØªØ±Ø¬Ø§Ø¹ Ø§Ù„Ù…Ø§Ù„ÙŠ (Ù…Ø®ØªØµØ±):\n"
    "â€¢ Ø®Ù„Ø§Ù„ Ø§Ù„Ø³Ø­Ø¨ ÙˆØ§Ù„Ø¥Ø¶Ø§ÙØ©: 100%\n"
    "â€¢ Ø§Ù„Ø£Ø³Ø¨ÙˆØ¹ Ø§Ù„Ø«Ø§Ù†ÙŠ: 70% (W)\n"
    "â€¢ Ø§Ù„Ø£Ø³Ø¨ÙˆØ¹ Ø§Ù„Ø«Ø§Ù„Ø«: 50% (W)\n"
    "â€¢ Ø¨Ø¹Ø¯ Ø§Ù„Ø£Ø³Ø¨ÙˆØ¹ Ø§Ù„Ø«Ø§Ù„Ø«: Ù„Ø§ ÙŠÙˆØ¬Ø¯ Ø§Ø³ØªØ±Ø¬Ø§Ø¹ (W)"
)

CONTACT_INFO = (
    "ğŸ“ Ø§Ù„ØªÙˆØ§ØµÙ„ â€“ Ø§Ù„Ø¬Ø§Ù…Ø¹Ø© Ø§Ù„Ø¹Ø±Ø¨ÙŠØ© Ø§Ù„Ù…ÙØªÙˆØ­Ø© (Ø§Ù„ÙƒÙˆÙŠØª)\n\n"
    "ğŸ“ Ø§Ù„Ø¹Ø§Ø±Ø¶ÙŠØ© Ø§Ù„ØµÙ†Ø§Ø¹ÙŠØ© â€“ Ø§Ù„ÙØ±ÙˆØ§Ù†ÙŠØ© â€“ Ø§Ù„ØµÙØ§Ø© 13033 â€“ Ø§Ù„ÙƒÙˆÙŠØª\n"
    "â˜ï¸ +965 24394400\n"
    "ğŸ“  +965 24394200\n"
    "âœ‰ï¸ info@aou.edu.kw\n\n"
    f"ğŸ”— Ù…ÙˆÙ‚Ø¹ Ø§Ù„Ø¬Ø§Ù…Ø¹Ø©: {UNIV_URL}\n"
    f"ğŸ“² Ù…Ù†ØµØ§Øª Ø§Ù„Ø¬Ø§Ù…Ø¹Ø©: {AOU_SOCIAL_LINKTREE}"
)

# =========================================================
# VALIDATION REGEX
# =========================================================
URL_REGEX = re.compile(r"(https?://\S+|www\.\S+|t\.me/\S+|chat\.whatsapp\.com/\S+|wa\.me/\S+|api\.whatsapp\.com/\S+)", flags=re.IGNORECASE)
TG_USERNAME_REGEX = re.compile(r"^@[A-Za-z0-9_]{4,}$")
TG_LINK_REGEX = re.compile(r"^(?:https?://)?t\.me/[A-Za-z0-9_+/=-]+", flags=re.IGNORECASE)
WA_LINK_REGEX = re.compile(r"^(?:https?://)?(?:chat\.whatsapp\.com/|wa\.me/|api\.whatsapp\.com/)\S+$", flags=re.IGNORECASE)

# =========================================================
# FIXED NAV BUTTONS
# =========================================================
BTN_BACK = "â¬…ï¸ Ø±Ø¬ÙˆØ¹"
BTN_HOME = "ğŸ  Ø§Ù„Ù‚Ø§Ø¦Ù…Ø© Ø§Ù„Ø±Ø¦ÙŠØ³ÙŠØ©"
BTN_CANCEL = "âŒ Ø¥Ù„ØºØ§Ø¡"
BTN_YES = "âœ… Ù†Ø¹Ù…"
BTN_NO = "âŒ Ù„Ø§"

BTN_ADMIN_SETTINGS = "âš™ï¸ Ø§Ù„Ø¥Ø¹Ø¯Ø§Ø¯Ø§Øª"

BTN_UNIV_NUMBERS = "â˜ï¸ Ø£Ø±Ù‚Ø§Ù… Ø§Ù„Ø¬Ø§Ù…Ø¹Ø©"
BTN_CONTACT_ADMIN = "ğŸ“© Ù…Ø±Ø§Ø³Ù„Ø© Ø§Ù„Ø¥Ø¯Ø§Ø±Ø©"
BTN_SOCIALS = "ğŸ“² Ù…Ù†ØµØ§Øª Ø§Ù„Ø¬Ø§Ù…Ø¹Ø©"

BTN_INBOX_SHOW = "ğŸ“¨ Ø±Ø³Ø§Ø¦Ù„ Ø§Ù„Ø£Ø¹Ø¶Ø§Ø¡"
BTN_INBOX_CLEAR = "ğŸ§¹ Ù…Ø³Ø­ Ø§Ù„Ø±Ø³Ø§Ø¦Ù„"

BTN_CAL_SHOW = "ğŸ“… Ø¹Ø±Ø¶ Ø§Ù„ØªÙ‚ÙˆÙŠÙ…"
BTN_CAL_REFRESH = "ğŸ”„ ØªØ­Ø¯ÙŠØ« Ø±ÙˆØ§Ø¨Ø· Ø§Ù„ØªÙ‚ÙˆÙŠÙ…"
BTN_CAL_AUTO_ON = "âœ… ØªØ´ØºÙŠÙ„ Ø§Ù„ØªØ­Ø¯ÙŠØ« Ø§Ù„ØªÙ„Ù‚Ø§Ø¦ÙŠ"
BTN_CAL_AUTO_OFF = "â›” Ø¥ÙŠÙ‚Ø§Ù Ø§Ù„ØªØ­Ø¯ÙŠØ« Ø§Ù„ØªÙ„Ù‚Ø§Ø¦ÙŠ"
BTN_CAL_SET_MANUAL = "âœï¸ ØªÙ‚ÙˆÙŠÙ… Ù…Ø®ØµØµ"
BTN_CAL_USE_AUTO = "ğŸŸ¢ Ø§Ø³ØªØ®Ø¯Ø§Ù… ØªÙ„Ù‚Ø§Ø¦ÙŠ"
BTN_CAL_CLEAR = "ğŸ—‘ï¸ Ø­Ø°Ù/Ø¥Ø®ÙØ§Ø¡ Ø§Ù„ØªÙ‚ÙˆÙŠÙ…"

BTN_GG_MENU = "ğŸ‘¥ Ø¥Ø¯Ø§Ø±Ø© Ø§Ù„Ù‚Ø±ÙˆØ¨ Ø§Ù„Ø¹Ø§Ù…"
BTN_CC_MENU = "ğŸ“² Ø®Ø¯Ù…Ø© Ø§Ù„Ø£Ø±Ù‚Ø§Ù…"
BTN_AM_MENU = "ğŸ‘® Ø¥Ø¯Ø§Ø±Ø© Ø§Ù„Ø£Ø¯Ù…Ù†"
BTN_HIDE_MENU = "ğŸ‘ï¸â€ğŸ—¨ï¸ Ø¥Ø®ÙØ§Ø¡/Ø¥Ø¸Ù‡Ø§Ø± Ø§Ù„Ø£Ø²Ø±Ø§Ø±"
BTN_RENAME_MENU = "âœï¸ ØªØ¹Ø¯ÙŠÙ„ Ø£Ø³Ù…Ø§Ø¡ Ø§Ù„Ø£Ø²Ø±Ø§Ø±"

BTN_GG_ADD = "â• Ø¥Ø¶Ø§ÙØ© Ù‚Ø±ÙˆØ¨"
BTN_GG_EDIT = "âœï¸ ØªØ¹Ø¯ÙŠÙ„ Ù‚Ø±ÙˆØ¨"
BTN_GG_DEL = "â– Ø­Ø°Ù Ù‚Ø±ÙˆØ¨"
BTN_GG_LIST = "ğŸ“‹ Ø¹Ø±Ø¶ Ø§Ù„Ù‚Ø±ÙˆØ¨Ø§Øª"
BTN_GG_USER_TG = "ğŸ“¢ Ù‚Ø±ÙˆØ¨Ø§Øª ØªÙŠÙ„ÙŠØ¬Ø±Ø§Ù…"
BTN_GG_USER_WA = "ğŸ“± Ù‚Ø±ÙˆØ¨Ø§Øª ÙˆØ§ØªØ³Ø§Ø¨"

BTN_CC_ENABLE = "âœ… ØªØ´ØºÙŠÙ„ Ø§Ù„Ø®Ø¯Ù…Ø©"
BTN_CC_DISABLE = "â›” Ø¥ÙŠÙ‚Ø§Ù Ø§Ù„Ø®Ø¯Ù…Ø©"
BTN_CC_SHOW_OK = "ğŸ“— Ø§Ù„Ø£Ø±Ù‚Ø§Ù… Ø§Ù„Ù…Ù‚Ø¨ÙˆÙ„Ø©"
BTN_CC_SHOW_BAD = "ğŸ“• Ø§Ù„Ø£Ø±Ù‚Ø§Ù… Ø§Ù„Ù…Ø±ÙÙˆØ¶Ø©"
BTN_CC_CLEAR_OK = "ğŸ§¹ Ù…Ø³Ø­ Ø§Ù„Ù…Ù‚Ø¨ÙˆÙ„Ø©"
BTN_CC_CLEAR_BAD = "ğŸ§¹ Ù…Ø³Ø­ Ø§Ù„Ù…Ø±ÙÙˆØ¶Ø©"
BTN_CC_EXPORT_XLSX = "ğŸ“¤ ØªØµØ¯ÙŠØ± Ø§Ù„Ø£Ø±Ù‚Ø§Ù… (Excel)"

BTN_AM_ADD = "â• Ø¥Ø¶Ø§ÙØ© Ø£Ø¯Ù…Ù†"
BTN_AM_DEL = "â– Ø­Ø°Ù Ø£Ø¯Ù…Ù†"
BTN_AM_LIST = "ğŸ“‹ Ø¹Ø±Ø¶ Ø§Ù„Ø£Ø¯Ù…Ù†"

BTN_UPLOAD_FILE = "ğŸ“ Ø±ÙØ¹ Ù…Ù„Ù/ØµÙˆØ±Ø©"
BTN_EDIT_TEXT = "âœï¸ ØªØ¹Ø¯ÙŠÙ„ Ø§Ù„Ù†Øµ"
BTN_DEL_FILE = "ğŸ—‘ï¸ Ø­Ø°Ù Ø§Ù„Ù…Ù„Ù"
BTN_DELETE_SECTION = "ğŸ—‘ï¸ Ø­Ø°Ù Ø§Ù„Ù‚Ø³Ù…"

# Colleges view
BTN_COLLEGE_ABOUT = "ğŸ“Œ Ù†Ø¨Ø°Ø©"
BTN_COLLEGE_URL = "ğŸ”— Ø±Ø§Ø¨Ø· Ø§Ù„ÙƒÙ„ÙŠØ©"
BTN_COLLEGE_WA = "ğŸ“± Ù‚Ø±ÙˆØ¨Ø§Øª ÙˆØ§ØªØ³Ø§Ø¨"
BTN_COLLEGE_TG = "ğŸ“¢ Ù‚Ø±ÙˆØ¨Ø§Øª ØªÙŠÙ„ÙŠØ¬Ø±Ø§Ù…"
BTN_ADD_WA = "â• Ø¥Ø¶Ø§ÙØ© ÙˆØ§ØªØ³Ø§Ø¨"
BTN_DEL_WA = "â– Ø­Ø°Ù ÙˆØ§ØªØ³Ø§Ø¨"
BTN_ADD_TG = "â• Ø¥Ø¶Ø§ÙØ© ØªÙŠÙ„ÙŠØ¬Ø±Ø§Ù…"
BTN_DEL_TG = "â– Ø­Ø°Ù ØªÙŠÙ„ÙŠØ¬Ø±Ø§Ù…"

# =========================================================
# DYNAMIC BUTTON KEYS (rename/hide)
# =========================================================
KEY_MAIN_COLLEGES = "main_colleges"
KEY_MAIN_ADMISSION = "main_admission"
KEY_MAIN_MAJORS = "main_majors"
KEY_MAIN_ADD_DROP = "main_add_drop"
KEY_MAIN_CALENDAR = "main_calendar"
KEY_MAIN_SUMMARIES = "main_summaries"
KEY_MAIN_SERVICES = "main_services"
KEY_MAIN_GROUPS = "main_groups"
KEY_MAIN_CONTACT = "main_contact"
KEY_MAIN_ABOUT = "main_about"

KEY_SERV_SCHEDULE = "serv_schedule"
KEY_SERV_ANNUAL = "serv_annual"
KEY_SERV_EXAMS = "serv_exams"
KEY_SERV_REG_CONT = "serv_reg_cont"
KEY_SERV_REG_NEW = "serv_reg_new"
KEY_SERV_ABSENCE = "serv_absence"
KEY_SERV_DEPRIVATION = "serv_deprivation"
KEY_SERV_STRIKE = "serv_strike"

KEY_SUM_BOOKS = "sum_books"
KEY_SUM_NOTES = "sum_notes"

DEFAULT_LABELS: Dict[str, str] = {
    KEY_MAIN_COLLEGES: "ğŸ« Ø§Ù„ÙƒÙ„ÙŠØ§Øª",
    KEY_MAIN_ADMISSION: "âœ… Ø§Ù„Ù‚Ø¨ÙˆÙ„",
    KEY_MAIN_MAJORS: "ğŸ“š Ø§Ù„ØªØ®ØµØµØ§Øª",
    KEY_MAIN_ADD_DROP: "âœï¸ Ø§Ù„Ø³Ø­Ø¨ ÙˆØ§Ù„Ø¥Ø¶Ø§ÙØ©",
    KEY_MAIN_CALENDAR: "ğŸ“… Ø§Ù„ØªÙ‚ÙˆÙŠÙ… Ø§Ù„Ø¬Ø§Ù…Ø¹ÙŠ",
    KEY_MAIN_SUMMARIES: "ğŸ—‚ï¸ Ø§Ù„Ù…Ù„Ø®ØµØ§Øª",
    KEY_MAIN_SERVICES: "ğŸ§° Ø§Ù„Ø®Ø¯Ù…Ø§Øª Ø§Ù„Ø·Ù„Ø§Ø¨ÙŠØ©",
    KEY_MAIN_GROUPS: "ğŸ‘¥ Ø§Ù„Ù‚Ø±ÙˆØ¨ Ø§Ù„Ø¹Ø§Ù…",
    KEY_MAIN_CONTACT: "ğŸ“ Ø§Ù„ØªÙˆØ§ØµÙ„",
    KEY_MAIN_ABOUT: "â„¹ï¸ Ø¹Ù† Ø§Ù„Ø¬Ø§Ù…Ø¹Ø©",
    KEY_SERV_SCHEDULE: "ğŸ—“ï¸ Ø§Ù„Ø¬Ø¯ÙˆÙ„",
    KEY_SERV_ANNUAL: "ğŸ—ºï¸ Ø§Ù„Ø®Ø·Ø© Ø§Ù„Ø³Ù†ÙˆÙŠØ©",
    KEY_SERV_EXAMS: "ğŸ§¾ Ø¬Ø¯ÙˆÙ„ Ø§Ù„Ø§Ù…ØªØ­Ø§Ù†Ø§Øª",
    KEY_SERV_REG_CONT: "ğŸ“ ØªÙ‚ÙˆÙŠÙ… ØªØ³Ø¬ÙŠÙ„ Ø§Ù„Ù…Ø³ØªÙ…Ø±ÙŠÙ†",
    KEY_SERV_REG_NEW: "ğŸ†• ØªÙ‚ÙˆÙŠÙ… ØªØ³Ø¬ÙŠÙ„ Ø§Ù„Ù…Ø³ØªØ¬Ø¯ÙŠÙ†",
    KEY_SERV_ABSENCE: "ğŸ“‰ Ù†Ø³Ø¨Ø© Ø§Ù„ØºÙŠØ§Ø¨",
    KEY_SERV_DEPRIVATION: "â›” Ø§Ù„Ø­Ø±Ù…Ø§Ù†",
    KEY_SERV_STRIKE: "ğŸ§¾ Ø·ÙŠ Ø§Ù„Ù‚ÙŠØ¯",
    KEY_SUM_BOOKS: "ğŸ“š ÙƒØªØ¨",
    KEY_SUM_NOTES: "ğŸ—‚ï¸ Ù…Ù„Ø®ØµØ§Øª",
}

HIDEABLE_KEYS: List[str] = list(DEFAULT_LABELS.keys())

# =========================================================
# CONTENT ITEMS (services/summaries + NEW: add_drop + calendar_attach)
# =========================================================
CONTENT_KEYS: Dict[str, str] = {
    KEY_SERV_SCHEDULE: "schedule",
    KEY_SERV_ANNUAL: "annual_plan",
    KEY_SERV_EXAMS: "exams",
    KEY_SERV_REG_CONT: "reg_cont",
    KEY_SERV_REG_NEW: "reg_new",
    KEY_SERV_ABSENCE: "absence",
    KEY_SERV_DEPRIVATION: "deprivation",
    KEY_SERV_STRIKE: "strike",
    KEY_SUM_BOOKS: "sum_books",
    KEY_SUM_NOTES: "sum_notes",
}
CONTENT_KEY_ADD_DROP = "add_drop"          # âœ… Ù…Ø­ØªÙˆÙ‰ Ø²Ø± Ø§Ù„Ø³Ø­Ø¨ ÙˆØ§Ù„Ø¥Ø¶Ø§ÙØ©
CONTENT_KEY_CALENDAR_ATTACH = "calendar_attach"  # âœ… Ù…Ø­ØªÙˆÙ‰ Ø²Ø± Ø§Ù„ØªÙ‚ÙˆÙŠÙ… (Ù…Ø±ÙÙ‚/Ù†Øµ)

# =========================================================
# DATA MODEL + STORAGE
# =========================================================
@dataclass
class BotData:
    users: Set[int]
    extra_admins: Set[int]
    inbox: List[Dict[str, str]]

    general_groups: List[Dict[str, str]]  # {id,name,telegram,whatsapp}
    colleges: Dict[str, Dict[str, Any]]  # {name:{about,url,whatsapp:list,telegram:list}}

    calendar: Dict[str, str]  # links only + manual text (legacy) + modes
    contact_collection_enabled: bool
    contacts_ok: List[Dict[str, str]]
    contacts_rejected: List[Dict[str, str]]

    content_items: Dict[str, Dict[str, str]]  # item_key->{text,file_id,file_type,file_name,mime_type,updated_at}
    hidden_buttons: List[str]
    button_labels: Dict[str, str]


def now_str() -> str:
    return time.strftime("%Y-%m-%d %H:%M:%S")


def clip(text: str) -> str:
    s = "" if text is None else str(text)
    return s if len(s) <= MAX_TEXT_LEN else s[: MAX_TEXT_LEN - 3] + "..."


def normalize_digits(s: str) -> str:
    trans = str.maketrans("Ù Ù¡Ù¢Ù£Ù¤Ù¥Ù¦Ù§Ù¨Ù©Û°Û±Û²Û³Û´ÛµÛ¶Û·Û¸Û¹", "01234567890123456789")
    return (s or "").translate(trans)


def contains_link(text: str) -> bool:
    return bool(text and URL_REGEX.search(text))


def safe_int(s: Any) -> Optional[int]:
    s = str(s).strip()
    return int(s) if s.isdigit() else None


def deny_no_perm() -> str:
    return "â›” Ù„ÙŠØ³ Ù„Ø¯ÙŠÙƒ ØµÙ„Ø§Ø­ÙŠØ© ÙÙŠ Ø§Ù„ØªØ¹Ø¯ÙŠÙ„."


def _default_calendar() -> Dict[str, str]:
    return {
        "pdf1": "",
        "pdf2": "",
        "last_updated": "Ù„Ù… ÙŠØªÙ… Ø§Ù„ØªØ­Ø¯ÙŠØ« Ø¨Ø¹Ø¯",
        "last_source": "ØºÙŠØ± Ù…Ø¹Ø±ÙˆÙ",  # website/manual/cleared
        "auto_enabled": "true",
        "display_mode": "auto",  # auto/manual/cleared
        "manual_text": "",  # legacy manual text (super admin)
    }


def _default_colleges() -> Dict[str, Dict[str, Any]]:
    base_url = "https://www.aou.edu.kw/ar/academic-programs/Pages/default.aspx"
    return {
        "ğŸ¢ ÙƒÙ„ÙŠØ© Ø¥Ø¯Ø§Ø±Ø© Ø§Ù„Ø£Ø¹Ù…Ø§Ù„": {"about": "Ù…Ø¬Ø§Ù„Ø§Øª Ø§Ù„Ø¥Ø¯Ø§Ø±Ø© ÙˆØ§Ù„Ù…Ø­Ø§Ø³Ø¨Ø© ÙˆØ§Ù„Ø§Ù‚ØªØµØ§Ø¯ ÙˆMIS ÙˆØ§Ù„ØªØ³ÙˆÙŠÙ‚.", "url": base_url, "whatsapp": [], "telegram": []},
        "ğŸ’» ÙƒÙ„ÙŠØ© Ø¯Ø±Ø§Ø³Ø§Øª Ø§Ù„Ø­Ø§Ø³ÙˆØ¨": {"about": "ØªÙ‚Ù†ÙŠØ© Ø§Ù„Ù…Ø¹Ù„ÙˆÙ…Ø§Øª ÙˆØ§Ù„Ù…Ù‡Ø§Ø±Ø§Øª Ø§Ù„Ø¨Ø±Ù…Ø¬ÙŠØ© ÙˆØ§Ù„Ø¯Ø¹Ù… Ø§Ù„ØªÙ‚Ù†ÙŠ.", "url": base_url, "whatsapp": [], "telegram": []},
        "ğŸ‘©â€ğŸ« ÙƒÙ„ÙŠØ© Ø§Ù„ØªØ±Ø¨ÙŠØ©": {"about": "Ø¨Ø±Ø§Ù…Ø¬ Ø§Ù„ØªØ±Ø¨ÙŠØ© ÙˆØ§Ù„ØªØ¹Ù„ÙŠÙ… ÙˆØ§Ù„ØªØ±Ø¨ÙŠØ© Ø§Ù„Ø®Ø§ØµØ© ÙˆØ§Ù„Ø§Ø¨ØªØ¯Ø§Ø¦ÙŠØ©.", "url": base_url, "whatsapp": [], "telegram": []},
        "ğŸŒ ÙƒÙ„ÙŠØ© Ø§Ù„Ù„ØºØ§Øª": {"about": "Ø§Ù„Ù„ØºØ§Øª ÙˆØ§Ù„ØªØ±Ø¬Ù…Ø© Ø­Ø³Ø¨ Ø§Ù„Ø¨Ø±Ø§Ù…Ø¬ Ø§Ù„Ù…ØªØ§Ø­Ø© ÙÙŠ Ø§Ù„ÙØ±Ø¹.", "url": base_url, "whatsapp": [], "telegram": []},
    }


def _default_content_items() -> Dict[str, Dict[str, str]]:
    def item() -> Dict[str, str]:
        return {"text": "", "file_id": "", "file_type": "", "file_name": "", "mime_type": "", "updated_at": ""}

    keys = set(CONTENT_KEYS.values())
    # âœ… Ø¥Ø¶Ø§ÙØ§Øª Ø¬Ø¯ÙŠØ¯Ø©:
    keys.add(CONTENT_KEY_ADD_DROP)
    keys.add(CONTENT_KEY_CALENDAR_ATTACH)
    return {k: item() for k in keys}


def _migrate_general_groups(raw: dict) -> List[Dict[str, str]]:
    gg_list = raw.get("general_groups", None)
    if isinstance(gg_list, list):
        out: List[Dict[str, str]] = []
        for item in gg_list:
            if not isinstance(item, dict):
                continue
            out.append(
                {
                    "id": str(item.get("id") or str(int(time.time() * 1000))),
                    "name": str(item.get("name") or "Ù‚Ø±ÙˆØ¨"),
                    "telegram": str(item.get("telegram") or ""),
                    "whatsapp": str(item.get("whatsapp") or ""),
                }
            )
        return out

    old_single = raw.get("general_group", "")
    if isinstance(old_single, str) and old_single.strip():
        return [{"id": str(int(time.time() * 1000)), "name": "Ø§Ù„Ù‚Ø±ÙˆØ¨ Ø§Ù„Ø¹Ø§Ù…", "telegram": old_single.strip(), "whatsapp": ""}]
    return []


def load_data() -> BotData:
    if not os.path.exists(DATA_FILE):
        return BotData(
            users=set(),
            extra_admins=set(),
            inbox=[],
            general_groups=[],
            colleges=_default_colleges(),
            calendar=_default_calendar(),
            contact_collection_enabled=False,
            contacts_ok=[],
            contacts_rejected=[],
            content_items=_default_content_items(),
            hidden_buttons=[],
            button_labels={},
        )

    try:
        with open(DATA_FILE, "r", encoding="utf-8") as f:
            raw = json.load(f)

        users = set(int(x) for x in raw.get("users", []) if str(x).isdigit())
        extra_admins = set(int(x) for x in raw.get("extra_admins", []) if str(x).isdigit())

        inbox = raw.get("inbox", [])
        if not isinstance(inbox, list):
            inbox = []

        general_groups = _migrate_general_groups(raw)

        colleges = raw.get("colleges", None)
        if not isinstance(colleges, dict) or not colleges:
            colleges = _default_colleges()
        else:
            for cname, cobj in list(colleges.items()):
                if not isinstance(cobj, dict):
                    colleges.pop(cname, None)
                    continue
                cobj.setdefault("about", "")
                cobj.setdefault("url", "")
                cobj.setdefault("whatsapp", [])
                cobj.setdefault("telegram", [])
                if not isinstance(cobj["whatsapp"], list):
                    cobj["whatsapp"] = []
                if not isinstance(cobj["telegram"], list):
                    cobj["telegram"] = []

        calendar = raw.get("calendar", {})
        if not isinstance(calendar, dict):
            calendar = _default_calendar()
        for k, v in _default_calendar().items():
            calendar.setdefault(k, v)

        contact_collection_enabled = bool(raw.get("contact_collection_enabled", False))

        contacts_ok = raw.get("contacts_ok", [])
        contacts_rejected = raw.get("contacts_rejected", [])
        if not isinstance(contacts_ok, list):
            contacts_ok = []
        if not isinstance(contacts_rejected, list):
            contacts_rejected = []

        defaults = _default_content_items()
        content_items = raw.get("content_items", None)
        if not isinstance(content_items, dict):
            content_items = defaults
        else:
            for k, v in defaults.items():
                if k not in content_items or not isinstance(content_items.get(k), dict):
                    content_items[k] = v
                else:
                    for fk, fv in v.items():
                        content_items[k].setdefault(fk, fv)

        hidden_buttons = raw.get("hidden_buttons", [])
        if not isinstance(hidden_buttons, list):
            hidden_buttons = []

        button_labels = raw.get("button_labels", {})
        if not isinstance(button_labels, dict):
            button_labels = {}

        return BotData(
            users=users,
            extra_admins=extra_admins,
            inbox=inbox,
            general_groups=general_groups,
            colleges=colleges,
            calendar={str(k): str(v) for k, v in calendar.items()},
            contact_collection_enabled=contact_collection_enabled,
            contacts_ok=contacts_ok,
            contacts_rejected=contacts_rejected,
            content_items={str(k): {str(kk): str(vv) for kk, vv in (v or {}).items()} for k, v in content_items.items()},
            hidden_buttons=[str(x) for x in hidden_buttons],
            button_labels={str(k): str(v) for k, v in button_labels.items()},
        )
    except Exception as e:
        LOG.exception("Load failed: %s", e)
        return BotData(
            users=set(),
            extra_admins=set(),
            inbox=[],
            general_groups=[],
            colleges=_default_colleges(),
            calendar=_default_calendar(),
            contact_collection_enabled=False,
            contacts_ok=[],
            contacts_rejected=[],
            content_items=_default_content_items(),
            hidden_buttons=[],
            button_labels={},
        )


def save_data(data: BotData) -> None:
    raw = {
        "users": list(data.users),
        "extra_admins": list(data.extra_admins),
        "inbox": data.inbox,
        "general_groups": data.general_groups,
        "colleges": data.colleges,
        "calendar": data.calendar,
        "contact_collection_enabled": data.contact_collection_enabled,
        "contacts_ok": data.contacts_ok,
        "contacts_rejected": data.contacts_rejected,
        "content_items": data.content_items,
        "hidden_buttons": data.hidden_buttons,
        "button_labels": data.button_labels,
    }
    tmp = f"{DATA_FILE}.tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(raw, f, ensure_ascii=False, indent=2)
    os.replace(tmp, DATA_FILE)


DATA = load_data()


def data_mutate(mutator) -> None:
    with DATA_LOCK:
        mutator()
        save_data(DATA)


def data_read(getter):
    with DATA_LOCK:
        return getter()


# =========================================================
# PERMISSIONS
# =========================================================
def is_super_admin(uid: int) -> bool:
    return uid in SUPER_ADMIN_IDS


def is_admin(uid: int) -> bool:
    with DATA_LOCK:
        return is_super_admin(uid) or (uid in DATA.extra_admins)


def all_admin_ids() -> Set[int]:
    with DATA_LOCK:
        return set(SUPER_ADMIN_IDS) | set(DATA.extra_admins)


# =========================================================
# BUTTON LABELS + HIDE
# =========================================================
def label_for(key: str) -> str:
    with DATA_LOCK:
        custom = DATA.button_labels.get(key, "").strip()
    return custom if custom else DEFAULT_LABELS.get(key, key)


def is_hidden_for_public(key: str) -> bool:
    with DATA_LOCK:
        return key in set(DATA.hidden_buttons)


def resolve_key_by_text(text: str, keys: List[str]) -> Optional[str]:
    for k in keys:
        if text == label_for(k):
            return k
    return None


def toggle_hidden(key: str) -> None:
    def _mut():
        s = set(DATA.hidden_buttons)
        if key in s:
            s.remove(key)
        else:
            s.add(key)
        DATA.hidden_buttons = sorted(s)

    data_mutate(_mut)


def set_label(key: str, new_label: str) -> None:
    new_label = (new_label or "").strip()

    def _mut():
        if not new_label:
            DATA.button_labels.pop(key, None)
        else:
            DATA.button_labels[key] = new_label

    data_mutate(_mut)


# =========================================================
# CONTACT GATE (KUWAIT ONLY)
# =========================================================
def normalize_kw_phone(raw: str) -> Tuple[Optional[str], str]:
    """
    Accept:
      - +965XXXXXXXX
      - 965XXXXXXXX
      - local 8 digits starting with 2/5/6/9 -> +965
    """
    s = normalize_digits(raw or "")
    s = re.sub(r"[^\d+]", "", s)

    if not s:
        return None, "empty"

    if s.startswith("00"):
        s = "+" + s[2:]

    if s.startswith("+"):
        if s.startswith("+965") and len(s) == 12 and s[4:].isdigit():
            return s, "ok"
        return None, "non_kw_country_code"

    if s.startswith("965") and len(s) == 11 and s[3:].isdigit():
        return "+965" + s[3:], "ok"

    if len(s) == 8 and s.isdigit() and s[0] in {"2", "5", "6", "9"}:
        return "+965" + s, "ok"

    return None, "invalid_format"


def contact_kb() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        [[KeyboardButton("ğŸ“± Ù…Ø´Ø§Ø±ÙƒØ© Ø±Ù‚Ù…ÙŠ", request_contact=True)], [KeyboardButton(BTN_HOME)]],
        resize_keyboard=True,
    )


def _user_has_saved_phone(uid: int) -> bool:
    uid_s = str(uid)
    with DATA_LOCK:
        for x in DATA.contacts_ok:
            if str(x.get("id")) == uid_s:
                return True
    return False


def contact_gate_required(update: Update) -> bool:
    if not update.effective_user or not update.effective_chat:
        return False
    if update.effective_chat.type != ChatType.PRIVATE:
        return False
    uid = update.effective_user.id
    if is_admin(uid):
        return False
    enabled = data_read(lambda: bool(DATA.contact_collection_enabled))
    return enabled and (not _user_has_saved_phone(uid))


def _append_contact_ok(user, phone: str) -> None:
    def _mut():
        uid_s = str(user.id)
        for x in DATA.contacts_ok:
            if str(x.get("id")) == uid_s:
                return
        DATA.contacts_ok.append(
            {
                "time": now_str(),
                "id": uid_s,
                "name": user.full_name or "â€”",
                "username": f"@{user.username}" if user.username else "Ø¨Ø¯ÙˆÙ† Ù…Ø¹Ø±Ù",
                "phone": phone,
            }
        )

    data_mutate(_mut)


def _append_contact_rejected(user, raw: str, reason: str) -> None:
    data_mutate(
        lambda: DATA.contacts_rejected.append(
            {
                "time": now_str(),
                "id": str(user.id),
                "name": user.full_name or "â€”",
                "username": f"@{user.username}" if user.username else "Ø¨Ø¯ÙˆÙ† Ù…Ø¹Ø±Ù",
                "raw": raw,
                "reason": reason,
            }
        )
    )


def contacts_ok_text(limit: int = 200) -> str:
    ok_list = data_read(lambda: list(DATA.contacts_ok))
    if not ok_list:
        return "ğŸ“— Ù„Ø§ ØªÙˆØ¬Ø¯ Ø£Ø±Ù‚Ø§Ù… ÙƒÙˆÙŠØªÙŠØ© Ù…Ø­ÙÙˆØ¸Ø©."
    lines = [f"ğŸ“— Ø§Ù„Ø£Ø±Ù‚Ø§Ù… Ø§Ù„ÙƒÙˆÙŠØªÙŠØ© Ø§Ù„Ù…Ù‚Ø¨ÙˆÙ„Ø©: {len(ok_list)}", ""]
    for i, x in enumerate(reversed(ok_list[-limit:]), start=1):
        lines.append(f"{i}) {x.get('phone')} | {x.get('name')} | {x.get('username')} | ID:{x.get('id')} | {x.get('time')}")
    return clip("\n".join(lines))


def contacts_bad_text(limit: int = 200) -> str:
    bad_list = data_read(lambda: list(DATA.contacts_rejected))
    if not bad_list:
        return "ğŸ“• Ù„Ø§ ØªÙˆØ¬Ø¯ Ø£Ø±Ù‚Ø§Ù… Ù…Ø±ÙÙˆØ¶Ø©."
    lines = [f"ğŸ“• Ø§Ù„Ø£Ø±Ù‚Ø§Ù… Ø§Ù„Ù…Ø±ÙÙˆØ¶Ø©: {len(bad_list)}", ""]
    for i, x in enumerate(reversed(bad_list[-limit:]), start=1):
        lines.append(
            f"{i}) {x.get('raw')} | Ø³Ø¨Ø¨:{x.get('reason')} | {x.get('name')} | {x.get('username')} | ID:{x.get('id')} | {x.get('time')}"
        )
    return clip("\n".join(lines))


def build_contacts_excel(path: str) -> None:
    ok_list = data_read(lambda: list(DATA.contacts_ok))
    bad_list = data_read(lambda: list(DATA.contacts_rejected))

    wb = Workbook()
    ws_ok = wb.active
    ws_ok.title = "Accepted"
    ws_bad = wb.create_sheet("Rejected")

    ok_headers = ["time", "id", "name", "username", "phone"]
    bad_headers = ["time", "id", "name", "username", "raw", "reason"]

    ws_ok.append(ok_headers)
    for x in ok_list:
        ws_ok.append([x.get("time", ""), x.get("id", ""), x.get("name", ""), x.get("username", ""), x.get("phone", "")])

    ws_bad.append(bad_headers)
    for x in bad_list:
        ws_bad.append([x.get("time", ""), x.get("id", ""), x.get("name", ""), x.get("username", ""), x.get("raw", ""), x.get("reason", "")])

    # Auto-fit-ish columns
    for ws in (ws_ok, ws_bad):
        for col in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col)
            max_len = 10
            for row in range(1, ws.max_row + 1):
                v = ws.cell(row=row, column=col).value
                if v is not None:
                    max_len = max(max_len, len(str(v)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 45)

    wb.save(path)


# =========================================================
# LINK NORMALIZATION
# =========================================================
def normalize_telegram_link(v: str) -> Optional[str]:
    v = (v or "").strip()
    if not v:
        return None
    if TG_USERNAME_REGEX.match(v):
        return v
    if v.lower().startswith("t.me/"):
        v = "https://" + v
    if TG_LINK_REGEX.match(v):
        if not v.lower().startswith("http"):
            v = "https://" + v
        return v
    return None


def normalize_whatsapp_link(v: str) -> Optional[str]:
    v = (v or "").strip()
    if not v:
        return None
    if v.lower().startswith("wa.me/"):
        v = "https://" + v
    if WA_LINK_REGEX.match(v):
        if not v.lower().startswith("http"):
            v = "https://" + v
        return v
    return None


def open_url_for_display(v: str) -> Optional[str]:
    s = (v or "").strip()
    if not s:
        return None
    if s.startswith("@"):
        return f"https://t.me/{s[1:]}"
    if s.lower().startswith("t.me/"):
        return "https://" + s
    if s.lower().startswith("http"):
        return s
    return None


def open_buttons(telegram_link: str, whatsapp_link: str) -> Optional[InlineKeyboardMarkup]:
    buttons: List[List[InlineKeyboardButton]] = []
    t_url = open_url_for_display(telegram_link)
    w_url = open_url_for_display(whatsapp_link)
    row: List[InlineKeyboardButton] = []
    if t_url:
        row.append(InlineKeyboardButton("ÙØªØ­ ØªÙŠÙ„ÙŠØ¬Ø±Ø§Ù…", url=t_url))
    if w_url:
        row.append(InlineKeyboardButton("ÙØªØ­ ÙˆØ§ØªØ³Ø§Ø¨", url=w_url))
    if row:
        buttons.append(row)
    return InlineKeyboardMarkup(buttons) if buttons else None


# =========================================================
# CONTENT ITEMS (services/summaries + add_drop + calendar_attach)
# =========================================================
def get_content_item(item_key: str) -> Dict[str, str]:
    with DATA_LOCK:
        base = _default_content_items().get(item_key, {})
        v = DATA.content_items.get(item_key, base)
        return dict(v)


def set_content_text(item_key: str, value: str) -> None:
    value = (value or "").strip()

    def _mut():
        DATA.content_items.setdefault(item_key, _default_content_items().get(item_key, {}))
        DATA.content_items[item_key]["text"] = value
        DATA.content_items[item_key]["updated_at"] = now_str()

    data_mutate(_mut)


def set_content_file(item_key: str, file_id: str, file_type: str, file_name: str, mime_type: str) -> None:
    def _mut():
        DATA.content_items.setdefault(item_key, _default_content_items().get(item_key, {}))
        DATA.content_items[item_key]["file_id"] = file_id
        DATA.content_items[item_key]["file_type"] = file_type
        DATA.content_items[item_key]["file_name"] = file_name
        DATA.content_items[item_key]["mime_type"] = mime_type
        DATA.content_items[item_key]["updated_at"] = now_str()

    data_mutate(_mut)


def delete_content_file(item_key: str) -> None:
    def _mut():
        DATA.content_items.setdefault(item_key, _default_content_items().get(item_key, {}))
        DATA.content_items[item_key]["file_id"] = ""
        DATA.content_items[item_key]["file_type"] = ""
        DATA.content_items[item_key]["file_name"] = ""
        DATA.content_items[item_key]["mime_type"] = ""
        DATA.content_items[item_key]["updated_at"] = now_str()

    data_mutate(_mut)


def delete_content_section(item_key: str) -> None:
    def _mut():
        DATA.content_items[item_key] = _default_content_items().get(item_key, {})
        DATA.content_items[item_key]["updated_at"] = now_str()

    data_mutate(_mut)


def content_is_empty(item_key: str) -> bool:
    item = get_content_item(item_key)
    return (not (item.get("text") or "").strip()) and (not (item.get("file_id") or "").strip())


async def send_content_to_user(update: Update, item_key: str, fallback_text: Optional[str] = None) -> None:
    if not update.message:
        return
    item = get_content_item(item_key)
    text = (item.get("text") or "").strip()
    file_id = (item.get("file_id") or "").strip()
    file_type = (item.get("file_type") or "").strip()

    if file_id:
        try:
            if file_type == "photo":
                await update.message.reply_photo(photo=file_id, caption=clip(text) if text else None)
            else:
                await update.message.reply_document(document=file_id, caption=clip(text) if text else None)
            return
        except Exception:
            pass

    if text:
        await update.message.reply_text(clip(text))
        return

    await update.message.reply_text(clip(fallback_text) if fallback_text else "ØºÙŠØ± Ù…ØªÙˆÙØ± Ø­Ø§Ù„ÙŠÙ‹Ø§.")


# =========================================================
# CALENDAR (links only, no PDF reading) + admin attachment via content_items
# =========================================================
def _http_get(url: str, timeout: int = 35) -> bytes:
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0 (AOU-Bot)"})
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return resp.read()


def _normalize_pdf_url(u: str) -> str:
    u = u.strip()
    if u.startswith("http://") or u.startswith("https://"):
        return u
    if u.startswith("/"):
        return "https://www.aou.edu.kw" + u
    return "https://www.aou.edu.kw/" + u


def _extract_pdf_links(html_text: str) -> List[str]:
    found = set()
    for m in re.findall(r'href="([^"]+\.pdf)"', html_text, flags=re.IGNORECASE):
        found.add(_normalize_pdf_url(m))
    for m in re.findall(r'src="([^"]+\.pdf)"', html_text, flags=re.IGNORECASE):
        found.add(_normalize_pdf_url(m))
    for m in re.findall(r'(https?://[^\s"<>]+\.pdf)', html_text, flags=re.IGNORECASE):
        found.add(m)
    for m in re.findall(r'(/[^"\s<>]+\.pdf)', html_text, flags=re.IGNORECASE):
        found.add(_normalize_pdf_url(m))
    return list(found)


def _score_pdf_link(url: str) -> int:
    u = url.lower()
    score = 0
    if "academic" in u or "calendar" in u:
        score += 5
    if "students" in u:
        score += 2
    if "publishingimages" in u:
        score += 1
    if "admission" in u:
        score -= 2
    return score


def cal_auto_enabled() -> bool:
    return str(data_read(lambda: DATA.calendar.get("auto_enabled", "true"))).lower() == "true"


def cal_display_mode() -> str:
    return str(data_read(lambda: DATA.calendar.get("display_mode", "auto"))).lower()


def refresh_calendar_links(force: bool = False) -> Tuple[bool, str]:
    if not force and (not cal_auto_enabled() or cal_display_mode() != "auto"):
        return True, "â¸ï¸ Ø§Ù„ØªØ­Ø¯ÙŠØ« Ø§Ù„ØªÙ„Ù‚Ø§Ø¦ÙŠ Ù…ÙˆÙ‚ÙˆÙ Ø£Ùˆ Ø§Ù„ØªÙ‚ÙˆÙŠÙ… Ù„ÙŠØ³ ØªÙ„Ù‚Ø§Ø¦ÙŠÙ‹Ø§."

    try:
        html_en = _http_get(AOU_CALENDAR_PAGE_EN).decode("utf-8", errors="ignore")
        html_ar = _http_get(AOU_CALENDAR_PAGE_AR).decode("utf-8", errors="ignore")
        links = list({*(_extract_pdf_links(html_en) + _extract_pdf_links(html_ar))})
        links = [x for x in links if x.lower().endswith(".pdf")]
        if not links:
            return False, "âŒ Ù„Ù… Ø£Ø¬Ø¯ Ø±ÙˆØ§Ø¨Ø· PDF ÙÙŠ ØµÙØ­Ø© Ø§Ù„ØªÙ‚ÙˆÙŠÙ…."

        links = sorted(links, key=_score_pdf_link, reverse=True)
        pdf1 = links[0]
        pdf2 = links[1] if len(links) > 1 else ""

        def _mut():
            DATA.calendar["pdf1"] = pdf1
            DATA.calendar["pdf2"] = pdf2
            DATA.calendar["last_updated"] = now_str()
            DATA.calendar["last_source"] = "website"

        data_mutate(_mut)
        return True, "âœ… ØªÙ… ØªØ­Ø¯ÙŠØ« Ø±ÙˆØ§Ø¨Ø· Ø§Ù„ØªÙ‚ÙˆÙŠÙ… Ù…Ù† Ù…ÙˆÙ‚Ø¹ Ø§Ù„Ø¬Ø§Ù…Ø¹Ø©."
    except Exception as e:
        LOG.exception("Calendar update failed: %s", e)
        return False, "âŒ Ø­Ø¯Ø« Ø®Ø·Ø£ Ø£Ø«Ù†Ø§Ø¡ ØªØ­Ø¯ÙŠØ« Ø±ÙˆØ§Ø¨Ø· Ø§Ù„ØªÙ‚ÙˆÙŠÙ…."


def calendar_text_plain() -> str:
    c = data_read(lambda: dict(DATA.calendar))
    mode = str(c.get("display_mode", "auto")).lower()
    auto = "âœ… Ø´ØºØ§Ù„" if str(c.get("auto_enabled", "true")).lower() == "true" else "â¸ï¸ Ù…ÙˆÙ‚ÙˆÙ"

    if mode == "cleared":
        return clip(
            "ğŸ“… Ø§Ù„ØªÙ‚ÙˆÙŠÙ… Ø§Ù„Ø¬Ø§Ù…Ø¹ÙŠ\n\n"
            "âš ï¸ Ø§Ù„ØªÙ‚ÙˆÙŠÙ… ØºÙŠØ± Ù…ØªÙˆÙØ± Ø­Ø§Ù„ÙŠÙ‹Ø§.\n"
            "ÙŠØ±Ø¬Ù‰ Ù…Ø±Ø§Ø¬Ø¹Ø© Ø§Ù„Ø¥Ø¯Ø§Ø±Ø© Ø£Ùˆ Ø§Ù†ØªØ¸Ø§Ø± Ø§Ù„ØªØ­Ø¯ÙŠØ«.\n\n"
            f"ğŸ”— ØµÙØ­Ø© Ø§Ù„ØªÙ‚ÙˆÙŠÙ…: {AOU_CALENDAR_PAGE_EN}"
        )

    if mode == "manual":
        manual = (c.get("manual_text") or "").strip() or "âš ï¸ Ù„Ø§ ÙŠÙˆØ¬Ø¯ Ù†Øµ Ù…Ø®ØµØµ."
        return clip(
            "ğŸ“… Ø§Ù„ØªÙ‚ÙˆÙŠÙ… Ø§Ù„Ø¬Ø§Ù…Ø¹ÙŠ (Ù…Ø®ØµØµ)\n\n"
            f"{manual}\n\n"
            f"Ø¢Ø®Ø± ØªØ­Ø¯ÙŠØ«: {c.get('last_updated','â€”')}\n"
            f"Ø§Ù„ØªØ­Ø¯ÙŠØ« Ø§Ù„ØªÙ„Ù‚Ø§Ø¦ÙŠ: {auto}\n"
            f"ğŸ”— ØµÙØ­Ø© Ø§Ù„ØªÙ‚ÙˆÙŠÙ…: {AOU_CALENDAR_PAGE_EN}"
        )

    pdf1 = c.get("pdf1", "") or "ØºÙŠØ± Ù…ØªÙˆÙØ±"
    pdf2 = c.get("pdf2", "") or "ØºÙŠØ± Ù…ØªÙˆÙØ±"
    return clip(
        "ğŸ“… Ø§Ù„ØªÙ‚ÙˆÙŠÙ… Ø§Ù„Ø¬Ø§Ù…Ø¹ÙŠ\n\n"
        f"Ø§Ù„ØªØ­Ø¯ÙŠØ« Ø§Ù„ØªÙ„Ù‚Ø§Ø¦ÙŠ: {auto}\n"
        f"Ø¢Ø®Ø± ØªØ­Ø¯ÙŠØ«: {c.get('last_updated','â€”')}\n\n"
        "Ø±ÙˆØ§Ø¨Ø· PDF (Ù…Ù† Ù…ÙˆÙ‚Ø¹ Ø§Ù„Ø¬Ø§Ù…Ø¹Ø©):\n"
        f"1) {pdf1}\n"
        f"2) {pdf2}\n\n"
        f"ğŸ”— ØµÙØ­Ø© Ø§Ù„ØªÙ‚ÙˆÙŠÙ…: {AOU_CALENDAR_PAGE_EN}"
    )


def calendar_links_buttons() -> Optional[InlineKeyboardMarkup]:
    c = data_read(lambda: dict(DATA.calendar))
    if str(c.get("display_mode", "auto")).lower() != "auto":
        return None
    pdf1 = (c.get("pdf1") or "").strip()
    pdf2 = (c.get("pdf2") or "").strip()
    rows: List[List[InlineKeyboardButton]] = []
    row: List[InlineKeyboardButton] = []
    if pdf1.startswith("http"):
        row.append(InlineKeyboardButton("ÙØªØ­ PDF 1", url=pdf1))
    if pdf2.startswith("http"):
        row.append(InlineKeyboardButton("ÙØªØ­ PDF 2", url=pdf2))
    if row:
        rows.append(row)
    return InlineKeyboardMarkup(rows) if rows else None


# =========================================================
# INBOX
# =========================================================
def push_inbox(item: Dict[str, str]) -> None:
    def _mut():
        DATA.inbox.append(item)
        if len(DATA.inbox) > MAX_INBOX:
            DATA.inbox[:] = DATA.inbox[-MAX_INBOX:]

    data_mutate(_mut)


def clear_inbox() -> None:
    data_mutate(lambda: DATA.inbox.clear())


def format_inbox_plain(last_n: int = 50) -> str:
    inbox = data_read(lambda: list(DATA.inbox))
    if not inbox:
        return "ğŸ“¨ Ù„Ø§ ØªÙˆØ¬Ø¯ Ø±Ø³Ø§Ø¦Ù„ Ù…Ø­ÙÙˆØ¸Ø©."
    msgs = inbox[-last_n:]
    lines = ["ğŸ“¨ Ø¢Ø®Ø± Ø±Ø³Ø§Ø¦Ù„ Ø§Ù„Ø£Ø¹Ø¶Ø§Ø¡:", ""]
    for i, m in enumerate(reversed(msgs), start=1):
        lines.append(f"{i}) Ø§Ù„ÙˆÙ‚Øª: {m.get('time','â€”')}")
        lines.append(f"   Ø§Ù„Ø§Ø³Ù…: {m.get('name','â€”')}")
        lines.append(f"   Ø§Ù„Ù…Ø³ØªØ®Ø¯Ù…: {m.get('username','Ø¨Ø¯ÙˆÙ† Ù…Ø¹Ø±Ù')}")
        lines.append(f"   ID: {m.get('id','â€”')}")
        lines.append(f"   Ø§Ù„Ø±Ø³Ø§Ù„Ø©: {m.get('text','')}")
        lines.append("-" * 30)
    return clip("\n".join(lines))


async def forward_to_admins(app: Application, text: str) -> None:
    for aid in all_admin_ids():
        try:
            await app.bot.send_message(chat_id=aid, text=clip(text))
        except Exception:
            pass


# =========================================================
# KEYBOARDS
# =========================================================
def kb_confirm() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        [[KeyboardButton(BTN_YES), KeyboardButton(BTN_NO)], [KeyboardButton(BTN_BACK), KeyboardButton(BTN_HOME)]],
        resize_keyboard=True,
    )


def kb_wait_cancel() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup([[KeyboardButton(BTN_CANCEL)], [KeyboardButton(BTN_BACK), KeyboardButton(BTN_HOME)]], resize_keyboard=True)


def kb_main(uid: int) -> ReplyKeyboardMarkup:
    isadm = is_admin(uid)

    def show(key: str) -> bool:
        return isadm or (not is_hidden_for_public(key))

    rows: List[List[KeyboardButton]] = []

    r1: List[KeyboardButton] = []
    if show(KEY_MAIN_COLLEGES):
        r1.append(KeyboardButton(label_for(KEY_MAIN_COLLEGES)))
    if show(KEY_MAIN_ADMISSION):
        r1.append(KeyboardButton(label_for(KEY_MAIN_ADMISSION)))
    if r1:
        rows.append(r1)

    r2: List[KeyboardButton] = []
    if show(KEY_MAIN_MAJORS):
        r2.append(KeyboardButton(label_for(KEY_MAIN_MAJORS)))
    if show(KEY_MAIN_ADD_DROP):
        r2.append(KeyboardButton(label_for(KEY_MAIN_ADD_DROP)))
    if r2:
        rows.append(r2)

    r3: List[KeyboardButton] = []
    if show(KEY_MAIN_CALENDAR):
        r3.append(KeyboardButton(label_for(KEY_MAIN_CALENDAR)))
    if show(KEY_MAIN_SUMMARIES):
        r3.append(KeyboardButton(label_for(KEY_MAIN_SUMMARIES)))
    if r3:
        rows.append(r3)

    r4: List[KeyboardButton] = []
    if show(KEY_MAIN_SERVICES):
        r4.append(KeyboardButton(label_for(KEY_MAIN_SERVICES)))
    if show(KEY_MAIN_GROUPS):
        r4.append(KeyboardButton(label_for(KEY_MAIN_GROUPS)))
    if r4:
        rows.append(r4)

    r5: List[KeyboardButton] = []
    if show(KEY_MAIN_CONTACT):
        r5.append(KeyboardButton(label_for(KEY_MAIN_CONTACT)))
    if show(KEY_MAIN_ABOUT):
        r5.append(KeyboardButton(label_for(KEY_MAIN_ABOUT)))
    if r5:
        rows.append(r5)

    if isadm:
        rows.append([KeyboardButton(BTN_ADMIN_SETTINGS)])

    return ReplyKeyboardMarkup(rows, resize_keyboard=True)


def kb_services(uid: int) -> ReplyKeyboardMarkup:
    isadm = is_admin(uid)

    def show(key: str) -> bool:
        return isadm or (not is_hidden_for_public(key))

    rows: List[List[KeyboardButton]] = []

    a: List[KeyboardButton] = []
    if show(KEY_SERV_SCHEDULE):
        a.append(KeyboardButton(label_for(KEY_SERV_SCHEDULE)))
    if show(KEY_SERV_ANNUAL):
        a.append(KeyboardButton(label_for(KEY_SERV_ANNUAL)))
    if a:
        rows.append(a)

    b: List[KeyboardButton] = []
    if show(KEY_SERV_EXAMS):
        b.append(KeyboardButton(label_for(KEY_SERV_EXAMS)))
    if show(KEY_SERV_ABSENCE):
        b.append(KeyboardButton(label_for(KEY_SERV_ABSENCE)))
    if b:
        rows.append(b)

    c: List[KeyboardButton] = []
    if show(KEY_SERV_DEPRIVATION):
        c.append(KeyboardButton(label_for(KEY_SERV_DEPRIVATION)))
    if show(KEY_SERV_STRIKE):
        c.append(KeyboardButton(label_for(KEY_SERV_STRIKE)))
    if c:
        rows.append(c)

    d: List[KeyboardButton] = []
    if show(KEY_SERV_REG_CONT):
        d.append(KeyboardButton(label_for(KEY_SERV_REG_CONT)))
    if show(KEY_SERV_REG_NEW):
        d.append(KeyboardButton(label_for(KEY_SERV_REG_NEW)))
    if d:
        rows.append(d)

    rows.append([KeyboardButton(BTN_BACK), KeyboardButton(BTN_HOME)])
    return ReplyKeyboardMarkup(rows, resize_keyboard=True)


def kb_summaries(uid: int) -> ReplyKeyboardMarkup:
    isadm = is_admin(uid)

    def show(key: str) -> bool:
        return isadm or (not is_hidden_for_public(key))

    rows: List[List[KeyboardButton]] = []
    r: List[KeyboardButton] = []
    if show(KEY_SUM_BOOKS):
        r.append(KeyboardButton(label_for(KEY_SUM_BOOKS)))
    if show(KEY_SUM_NOTES):
        r.append(KeyboardButton(label_for(KEY_SUM_NOTES)))
    if r:
        rows.append(r)
    rows.append([KeyboardButton(BTN_BACK), KeyboardButton(BTN_HOME)])
    return ReplyKeyboardMarkup(rows, resize_keyboard=True)


def kb_contact(uid: int) -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        [
            [KeyboardButton(BTN_UNIV_NUMBERS), KeyboardButton(BTN_SOCIALS)],
            [KeyboardButton(BTN_CONTACT_ADMIN)],
            [KeyboardButton(BTN_BACK), KeyboardButton(BTN_HOME)],
        ],
        resize_keyboard=True,
    )


def kb_admin(uid: int) -> ReplyKeyboardMarkup:
    rows: List[List[KeyboardButton]] = [
        [KeyboardButton(BTN_INBOX_SHOW), KeyboardButton(BTN_INBOX_CLEAR)],
        [KeyboardButton(BTN_CAL_SHOW), KeyboardButton(BTN_CAL_REFRESH)],
        [KeyboardButton(BTN_GG_MENU), KeyboardButton(BTN_CC_MENU)],
        [KeyboardButton(BTN_BACK), KeyboardButton(BTN_HOME)],
    ]
    if is_super_admin(uid):
        auto_btn = BTN_CAL_AUTO_OFF if cal_auto_enabled() else BTN_CAL_AUTO_ON
        rows.insert(2, [KeyboardButton(auto_btn), KeyboardButton(BTN_CAL_SET_MANUAL)])
        rows.insert(3, [KeyboardButton(BTN_CAL_USE_AUTO), KeyboardButton(BTN_CAL_CLEAR)])
        rows.insert(4, [KeyboardButton(BTN_HIDE_MENU), KeyboardButton(BTN_RENAME_MENU)])
        rows.insert(5, [KeyboardButton(BTN_AM_MENU)])
    return ReplyKeyboardMarkup(rows, resize_keyboard=True)


def kb_general_groups_admin() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        [
            [KeyboardButton(BTN_GG_LIST)],
            [KeyboardButton(BTN_GG_ADD), KeyboardButton(BTN_GG_EDIT)],
            [KeyboardButton(BTN_GG_DEL)],
            [KeyboardButton(BTN_BACK), KeyboardButton(BTN_HOME)],
        ],
        resize_keyboard=True,
    )


def kb_general_groups_user(uid: int) -> ReplyKeyboardMarkup:
    rows = [[KeyboardButton(BTN_GG_USER_TG), KeyboardButton(BTN_GG_USER_WA)], [KeyboardButton(BTN_BACK), KeyboardButton(BTN_HOME)]]
    if is_admin(uid):
        rows.insert(1, [KeyboardButton(BTN_GG_MENU)])
    return ReplyKeyboardMarkup(rows, resize_keyboard=True)


def kb_contact_service_admin(uid: int) -> ReplyKeyboardMarkup:
    enabled = data_read(lambda: bool(DATA.contact_collection_enabled))
    # âœ… Ø¹Ø±Ø¶ Ù„Ù„Ø£Ø¯Ù…Ù†ØŒ Ù„ÙƒÙ† Ø§Ù„ØªØºÙŠÙŠØ±/Ø§Ù„Ù…Ø³Ø­ Ù„Ù„Ø³ÙˆØ¨Ø± ÙÙ‚Ø·
    rows: List[List[KeyboardButton]] = [
        [KeyboardButton(BTN_CC_SHOW_OK), KeyboardButton(BTN_CC_SHOW_BAD)],
        [KeyboardButton(BTN_CC_EXPORT_XLSX)],
    ]
    if is_super_admin(uid):
        status = BTN_CC_DISABLE if enabled else BTN_CC_ENABLE
        rows.insert(0, [KeyboardButton(status)])
        rows.append([KeyboardButton(BTN_CC_CLEAR_OK), KeyboardButton(BTN_CC_CLEAR_BAD)])
    rows.append([KeyboardButton(BTN_BACK), KeyboardButton(BTN_HOME)])
    return ReplyKeyboardMarkup(rows, resize_keyboard=True)


def kb_admin_manage() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        [
            [KeyboardButton(BTN_AM_LIST)],
            [KeyboardButton(BTN_AM_ADD), KeyboardButton(BTN_AM_DEL)],
            [KeyboardButton(BTN_BACK), KeyboardButton(BTN_HOME)],
        ],
        resize_keyboard=True,
    )


def kb_content_admin(parent_menu_text: str) -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        [
            [KeyboardButton(BTN_UPLOAD_FILE), KeyboardButton(BTN_DEL_FILE)],
            [KeyboardButton(BTN_EDIT_TEXT), KeyboardButton(BTN_DELETE_SECTION)],
            [KeyboardButton(BTN_BACK), KeyboardButton(BTN_HOME)],
            [KeyboardButton(parent_menu_text)],
        ],
        resize_keyboard=True,
    )


def kb_colleges(uid: int) -> ReplyKeyboardMarkup:
    names = data_read(lambda: list(DATA.colleges.keys()))
    rows: List[List[KeyboardButton]] = []
    for i in range(0, len(names), 2):
        row = [KeyboardButton(names[i])]
        if i + 1 < len(names):
            row.append(KeyboardButton(names[i + 1]))
        rows.append(row)
    rows.append([KeyboardButton(BTN_BACK), KeyboardButton(BTN_HOME)])
    return ReplyKeyboardMarkup(rows, resize_keyboard=True)


def kb_college_view(uid: int) -> ReplyKeyboardMarkup:
    rows = [
        [KeyboardButton(BTN_COLLEGE_ABOUT), KeyboardButton(BTN_COLLEGE_URL)],
        [KeyboardButton(BTN_COLLEGE_WA), KeyboardButton(BTN_COLLEGE_TG)],
    ]
    if is_admin(uid):
        rows += [
            [KeyboardButton(BTN_ADD_WA), KeyboardButton(BTN_DEL_WA)],
            [KeyboardButton(BTN_ADD_TG), KeyboardButton(BTN_DEL_TG)],
        ]
    rows.append([KeyboardButton(BTN_BACK), KeyboardButton(BTN_HOME)])
    return ReplyKeyboardMarkup(rows, resize_keyboard=True)


# =========================================================
# USER STATE / MODES
# =========================================================
USER_MODE = "mode"
USER_SELECTED_COLLEGE = "selected_college"
USER_SELECTED_ITEMKEY = "selected_itemkey"
USER_PARENT_MENU = "parent_menu"
USER_CONFIRM = "confirm"  # {action,payload,return_to}
USER_STEP = "step"
USER_TEMP = "temp"
USER_SELECTED_ID = "selected_id"
USER_CHOICE = "choice"
USER_RENAME_KEY = "rename_key"

MODE_NORMAL = "normal"
MODE_SUPPORT = "support"
MODE_SERVICES = "services"
MODE_SUMMARIES = "summaries"
MODE_CONTACT_MENU = "contact_menu"
MODE_COLLEGES = "colleges"
MODE_COLLEGE_VIEW = "college_view"
MODE_GG_USER_MENU = "gg_user_menu"

MODE_CONTENT_EDIT_TEXT = "content_edit_text"
MODE_CONTENT_UPLOAD_FILE = "content_upload_file"

MODE_CC_MENU = "cc_menu"
MODE_GG_MENU = "gg_menu"
MODE_GG_ADD = "gg_add"
MODE_GG_DEL = "gg_del"
MODE_GG_EDIT_SELECT = "gg_edit_select"
MODE_GG_EDIT_FIELD = "gg_edit_field"
MODE_GG_EDIT_VALUE = "gg_edit_value"

MODE_ADD_WA = "add_wa"
MODE_DEL_WA = "del_wa"
MODE_ADD_TG = "add_tg"
MODE_DEL_TG = "del_tg"

MODE_CAL_MANUAL_WAIT = "cal_manual_wait"
MODE_HIDE_MENU = "hide_menu"
MODE_RENAME_MENU = "rename_menu"
MODE_RENAME_VALUE = "rename_value"

MODE_AM_MENU = "am_menu"
MODE_AM_ADD = "am_add"
MODE_AM_DEL = "am_del"

MODE_CONFIRM = "confirm"


def set_mode(context: ContextTypes.DEFAULT_TYPE, mode: str) -> None:
    context.user_data[USER_MODE] = mode


def get_mode(context: ContextTypes.DEFAULT_TYPE) -> str:
    return context.user_data.get(USER_MODE, MODE_NORMAL)


def set_selected_college(context: ContextTypes.DEFAULT_TYPE, cname: Optional[str]) -> None:
    if cname:
        context.user_data[USER_SELECTED_COLLEGE] = cname
    else:
        context.user_data.pop(USER_SELECTED_COLLEGE, None)


def get_selected_college(context: ContextTypes.DEFAULT_TYPE) -> Optional[str]:
    return context.user_data.get(USER_SELECTED_COLLEGE)


def set_selected_content(context: ContextTypes.DEFAULT_TYPE, item_key: Optional[str], parent_menu_text: Optional[str]) -> None:
    if item_key:
        context.user_data[USER_SELECTED_ITEMKEY] = item_key
    else:
        context.user_data.pop(USER_SELECTED_ITEMKEY, None)
    if parent_menu_text:
        context.user_data[USER_PARENT_MENU] = parent_menu_text
    else:
        context.user_data.pop(USER_PARENT_MENU, None)


def get_selected_itemkey(context: ContextTypes.DEFAULT_TYPE) -> Optional[str]:
    return context.user_data.get(USER_SELECTED_ITEMKEY)


def get_parent_menu_text(context: ContextTypes.DEFAULT_TYPE) -> Optional[str]:
    return context.user_data.get(USER_PARENT_MENU)


def reset_flow(context: ContextTypes.DEFAULT_TYPE) -> None:
    context.user_data.clear()


# =========================================================
# CONFIRM SYSTEM
# =========================================================
def start_confirm(context: ContextTypes.DEFAULT_TYPE, action: str, payload: dict, return_to: dict) -> None:
    context.user_data[USER_CONFIRM] = {"action": action, "payload": payload, "return_to": return_to}
    set_mode(context, MODE_CONFIRM)


async def confirm_return(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not update.message or not update.effective_user:
        return
    uid = update.effective_user.id

    ret = (context.user_data.get(USER_CONFIRM) or {}).get("return_to", {}) or {}
    target = ret.get("target", "main")

    context.user_data.pop(USER_CONFIRM, None)
    context.user_data.pop(USER_STEP, None)
    context.user_data.pop(USER_TEMP, None)
    context.user_data.pop(USER_SELECTED_ID, None)
    context.user_data.pop(USER_CHOICE, None)
    context.user_data.pop(USER_RENAME_KEY, None)
    set_mode(context, MODE_NORMAL)

    if target == "admin":
        await update.message.reply_text("âš™ï¸", reply_markup=kb_admin(uid))
        return
    if target == "services":
        set_mode(context, MODE_SERVICES)
        await update.message.reply_text("ğŸ§°", reply_markup=kb_services(uid))
        return
    if target == "summaries":
        set_mode(context, MODE_SUMMARIES)
        await update.message.reply_text("ğŸ—‚ï¸", reply_markup=kb_summaries(uid))
        return
    if target == "gg_admin":
        set_mode(context, MODE_GG_MENU)
        await update.message.reply_text("ğŸ‘¥", reply_markup=kb_general_groups_admin())
        return
    if target == "college_view":
        await update.message.reply_text("ğŸ«", reply_markup=kb_college_view(uid))
        return
    if target == "am_menu":
        set_mode(context, MODE_AM_MENU)
        await update.message.reply_text("ğŸ‘®", reply_markup=kb_admin_manage())
        return

    await update.message.reply_text("ğŸ ", reply_markup=kb_main(uid))


async def run_confirm_action(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not update.message or not update.effective_user:
        return
    uid = update.effective_user.id

    cfg = context.user_data.get(USER_CONFIRM) or {}
    action = str(cfg.get("action") or "")
    payload: Dict[str, Any] = cfg.get("payload") or {}

    super_only = {
        "toggle_hide",
        "rename_button",
        "calendar_clear",
        "contact_toggle",
        "contact_clear_ok",
        "contact_clear_bad",
        "admin_add",
        "admin_del",
        "calendar_use_auto",
        "calendar_auto_on",
        "calendar_auto_off",
    }
    admin_required = {
        "inbox_clear",
        "content_del_file",
        "content_del_section",
        "gg_add",
        "gg_del",
        "college_add_wa",
        "college_del_wa",
        "college_add_tg",
        "college_del_tg",
    }

    if action in super_only and not is_super_admin(uid):
        await update.message.reply_text(deny_no_perm(), reply_markup=kb_admin(uid))
        await confirm_return(update, context)
        return

    if action in admin_required and not is_admin(uid):
        await update.message.reply_text("âŒ ØºÙŠØ± Ù…ØªØ§Ø­.", reply_markup=kb_main(uid))
        await confirm_return(update, context)
        return

    # ---- Execute ----
    if action == "inbox_clear":
        clear_inbox()

    elif action == "toggle_hide":
        key = str(payload.get("key", ""))
        if key in HIDEABLE_KEYS:
            toggle_hidden(key)

    elif action == "rename_button":
        key = str(payload.get("key", ""))
        value = str(payload.get("value", ""))
        if key in HIDEABLE_KEYS:
            set_label(key, value)

    elif action == "content_del_file":
        item_key = str(payload.get("item_key", ""))
        delete_content_file(item_key)

    elif action == "content_del_section":
        item_key = str(payload.get("item_key", ""))
        delete_content_section(item_key)

    elif action == "gg_add":
        group = payload.get("group")
        if isinstance(group, dict):
            def _mut():
                DATA.general_groups.append(
                    {
                        "id": str(group.get("id")),
                        "name": str(group.get("name", "Ù‚Ø±ÙˆØ¨")),
                        "telegram": str(group.get("telegram", "")),
                        "whatsapp": str(group.get("whatsapp", "")),
                    }
                )
            data_mutate(_mut)

    elif action == "gg_del":
        gid = str(payload.get("id", ""))
        def _mut():
            DATA.general_groups = [x for x in DATA.general_groups if str(x.get("id")) != gid]
        data_mutate(_mut)

    elif action == "contact_toggle":
        enabled = bool(payload.get("enabled"))
        data_mutate(lambda: setattr(DATA, "contact_collection_enabled", enabled))

    elif action == "contact_clear_ok":
        data_mutate(lambda: DATA.contacts_ok.clear())

    elif action == "contact_clear_bad":
        data_mutate(lambda: DATA.contacts_rejected.clear())

    elif action == "calendar_clear":
        def _mut():
            DATA.calendar["display_mode"] = "cleared"
            DATA.calendar["manual_text"] = ""
            DATA.calendar["pdf1"] = ""
            DATA.calendar["pdf2"] = ""
            DATA.calendar["last_updated"] = now_str()
            DATA.calendar["last_source"] = "cleared"
        data_mutate(_mut)

    elif action == "calendar_use_auto":
        def _mut():
            DATA.calendar["display_mode"] = "auto"
            DATA.calendar["last_updated"] = now_str()
            DATA.calendar["last_source"] = "website"
        data_mutate(_mut)

    elif action == "calendar_auto_on":
        data_mutate(lambda: DATA.calendar.__setitem__("auto_enabled", "true"))

    elif action == "calendar_auto_off":
        data_mutate(lambda: DATA.calendar.__setitem__("auto_enabled", "false"))

    elif action == "admin_add":
        new_id = safe_int(payload.get("id", ""))
        if new_id and new_id not in SUPER_ADMIN_IDS:
            data_mutate(lambda: DATA.extra_admins.add(new_id))

    elif action == "admin_del":
        rem_id = safe_int(payload.get("id", ""))
        if rem_id and rem_id not in SUPER_ADMIN_IDS:
            data_mutate(lambda: DATA.extra_admins.discard(rem_id))

    elif action == "college_add_wa":
        cname = str(payload.get("cname", ""))
        link = str(payload.get("link", ""))
        def _mut():
            col = DATA.colleges.get(cname)
            if isinstance(col, dict):
                lst = col.get("whatsapp", [])
                if isinstance(lst, list) and link not in lst:
                    lst.append(link)
                    col["whatsapp"] = lst
        data_mutate(_mut)

    elif action == "college_del_wa":
        cname = str(payload.get("cname", ""))
        link = str(payload.get("link", ""))
        def _mut():
            col = DATA.colleges.get(cname)
            if isinstance(col, dict):
                lst = col.get("whatsapp", [])
                if isinstance(lst, list):
                    col["whatsapp"] = [x for x in lst if x != link]
        data_mutate(_mut)

    elif action == "college_add_tg":
        cname = str(payload.get("cname", ""))
        link = str(payload.get("link", ""))
        def _mut():
            col = DATA.colleges.get(cname)
            if isinstance(col, dict):
                lst = col.get("telegram", [])
                if isinstance(lst, list) and link not in lst:
                    lst.append(link)
                    col["telegram"] = lst
        data_mutate(_mut)

    elif action == "college_del_tg":
        cname = str(payload.get("cname", ""))
        link = str(payload.get("link", ""))
        def _mut():
            col = DATA.colleges.get(cname)
            if isinstance(col, dict):
                lst = col.get("telegram", [])
                if isinstance(lst, list):
                    col["telegram"] = [x for x in lst if x != link]
        data_mutate(_mut)

    await update.message.reply_text("âœ… ØªÙ… Ø§Ù„ØªÙ†ÙÙŠØ°.", reply_markup=kb_main(uid))
    await confirm_return(update, context)


# =========================================================
# TELEGRAM COMMAND SCOPES
# =========================================================
async def refresh_scoped_commands(app: Application) -> None:
    admin_cmds = [
        BotCommand("start", "ØªØ´ØºÙŠÙ„ Ø§Ù„Ø¨ÙˆØª"),
        BotCommand("help", "Ù…Ø³Ø§Ø¹Ø¯Ø©"),
        BotCommand("myid", "Ø¹Ø±Ø¶ Ø¢ÙŠØ¯ÙŠÙƒ"),
        BotCommand("admin", "Ø¥Ø¹Ø¯Ø§Ø¯Ø§Øª (Ù„Ù„Ø£Ø¯Ù…Ù†)"),
    ]
    user_cmds = [c for c in admin_cmds if c.command != "admin"]

    try:
        await app.bot.set_my_commands(user_cmds, scope=BotCommandScopeDefault())
    except Exception:
        pass

    for aid in all_admin_ids():
        try:
            await app.bot.set_my_commands(admin_cmds, scope=BotCommandScopeChat(chat_id=aid))
        except BadRequest:
            pass
        except Exception:
            pass


async def post_init(app: Application) -> None:
    await refresh_scoped_commands(app)


# =========================================================
# AUTO JOBS
# =========================================================
async def calendar_auto_job(context: ContextTypes.DEFAULT_TYPE) -> None:
    try:
        ok, note = await asyncio.to_thread(refresh_calendar_links, False)
        LOG.info("Auto calendar => %s | %s", ok, note)
    except Exception as e:
        LOG.exception("Auto job error: %s", e)


def setup_jobs(app: Application) -> None:
    app.job_queue.run_repeating(calendar_auto_job, interval=12 * 60 * 60, first=90)


# =========================================================
# COMMANDS
# =========================================================
async def start_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not update.message or not update.effective_user:
        return
    uid = update.effective_user.id
    data_mutate(lambda: DATA.users.add(uid))
    reset_flow(context)

    if contact_gate_required(update):
        await update.message.reply_text(WELCOME_TEXT)
        await update.message.reply_text(CONTACT_PROMPT_TEXT)
        await update.message.reply_text(IMPORTANT_NOTICE_TEXT, reply_markup=contact_kb())
        return

    await update.message.reply_text(WELCOME_TEXT, reply_markup=kb_main(uid))


async def help_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not update.message or not update.effective_user:
        return
    uid = update.effective_user.id
    await update.message.reply_text("ğŸ†˜ Ø§Ù„Ù…Ø³Ø§Ø¹Ø¯Ø©\nØ§Ø³ØªØ®Ø¯Ù… Ø§Ù„Ø£Ø²Ø±Ø§Ø± Ù„Ù„ØªÙ†Ù‚Ù„.", reply_markup=kb_main(uid))


async def myid_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not update.message or not update.effective_user:
        return
    await update.message.reply_text(f"ğŸ†” Ø¢ÙŠØ¯ÙŠÙƒ: {update.effective_user.id}")


async def admin_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not update.message or not update.effective_user:
        return
    uid = update.effective_user.id
    if not is_admin(uid):
        await update.message.reply_text("âŒ ØºÙŠØ± Ù…ØªØ§Ø­.", reply_markup=kb_main(uid))
        return
    reset_flow(context)
    await update.message.reply_text("âš™ï¸ Ø¥Ø¹Ø¯Ø§Ø¯Ø§Øª:", reply_markup=kb_admin(uid))


# =========================================================
# CONTACT HANDLER
# =========================================================
async def handle_contact(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not update.message or not update.effective_user or not update.effective_chat:
        return
    if update.effective_chat.type != ChatType.PRIVATE:
        return

    uid = update.effective_user.id

    if is_admin(uid):
        await update.message.reply_text("âœ… Ø£Ù†Øª Ø£Ø¯Ù…Ù† â€” Ù„Ù† ÙŠØªÙ… Ø­ÙØ¸ Ø±Ù‚Ù…Ùƒ.", reply_markup=kb_main(uid))
        return

    enabled = data_read(lambda: bool(DATA.contact_collection_enabled))
    if not enabled:
        await update.message.reply_text("â›” Ø®Ø¯Ù…Ø© Ø§Ù„Ø£Ø±Ù‚Ø§Ù… ØºÙŠØ± Ù…ÙØ¹Ù„Ø© Ø­Ø§Ù„ÙŠÙ‹Ø§.", reply_markup=kb_main(uid))
        return

    if _user_has_saved_phone(uid):
        await update.message.reply_text("âœ… Ø±Ù‚Ù…Ùƒ Ù…Ø­ÙÙˆØ¸ Ù…Ø³Ø¨Ù‚Ù‹Ø§.", reply_markup=kb_main(uid))
        return

    c = update.message.contact
    if not c:
        return

    if c.user_id and int(c.user_id) != int(uid):
        raw = (c.phone_number or "").strip()
        _append_contact_rejected(update.effective_user, raw, "not_self")
        await update.message.reply_text("âŒ Ù„Ø§Ø²Ù… ØªØ´Ø§Ø±Ùƒ Ø±Ù‚Ù…Ùƒ Ø£Ù†Øª ÙÙ‚Ø·.", reply_markup=contact_kb())
        return

    raw = (c.phone_number or "").strip()
    e164, reason = normalize_kw_phone(raw)
    if e164:
        _append_contact_ok(update.effective_user, e164)
        await update.message.reply_text("âœ… ØªÙ… Ù‚Ø¨ÙˆÙ„Ùƒ ÙˆØ­ÙØ¸ Ø±Ù‚Ù…Ùƒ Ø¨Ù†Ø¬Ø§Ø­ ğŸ¤", reply_markup=kb_main(uid))
    else:
        _append_contact_rejected(update.effective_user, raw, reason)
        await update.message.reply_text("âŒ Ø§Ù„Ø±Ù‚Ù… ØºÙŠØ± ÙƒÙˆÙŠØªÙŠ Ø£Ùˆ ØºÙŠØ± ØµØ­ÙŠØ­. ØªÙ… Ø±ÙØ¶Ù‡.", reply_markup=contact_kb())


# =========================================================
# MEDIA HANDLER (admin upload for selected content)
# =========================================================
async def handle_media(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not update.message or not update.effective_user:
        return
    uid = update.effective_user.id
    if not is_admin(uid):
        return
    if get_mode(context) != MODE_CONTENT_UPLOAD_FILE:
        return

    item_key = get_selected_itemkey(context)
    parent_menu_text = get_parent_menu_text(context) or label_for(KEY_MAIN_SERVICES)
    if not item_key:
        set_mode(context, MODE_NORMAL)
        await update.message.reply_text("âŒ Ø®Ø·Ø£ ÙÙŠ Ø§Ù„Ø­Ø§Ù„Ø©.", reply_markup=kb_main(uid))
        return

    if update.message.photo:
        file_id = update.message.photo[-1].file_id
        set_content_file(item_key, file_id=file_id, file_type="photo", file_name="photo.jpg", mime_type="image/jpeg")
        set_mode(context, MODE_NORMAL)
        await update.message.reply_text("âœ… ØªÙ… Ø±ÙØ¹ Ø§Ù„ØµÙˆØ±Ø©.", reply_markup=kb_content_admin(parent_menu_text))
        return

    doc = update.message.document
    if doc:
        set_content_file(
            item_key,
            file_id=doc.file_id,
            file_type="document",
            file_name=doc.file_name or "file",
            mime_type=doc.mime_type or "",
        )
        set_mode(context, MODE_NORMAL)
        await update.message.reply_text("âœ… ØªÙ… Ø±ÙØ¹ Ø§Ù„Ù…Ù„Ù.", reply_markup=kb_content_admin(parent_menu_text))
        return

    await update.message.reply_text("âŒ Ø£Ø±Ø³Ù„ Ù…Ù„Ù Ø£Ùˆ ØµÙˆØ±Ø©.", reply_markup=kb_wait_cancel())


# =========================================================
# GENERAL GROUPS HELPERS
# =========================================================
def _gg_find(gg_id: str) -> Optional[Dict[str, str]]:
    with DATA_LOCK:
        for g in DATA.general_groups:
            if str(g.get("id")) == str(gg_id):
                return dict(g)
    return None


def _gg_list_text() -> str:
    groups = data_read(lambda: list(DATA.general_groups))
    if not groups:
        return "âŒ Ù„Ø§ ØªÙˆØ¬Ø¯ Ù‚Ø±ÙˆØ¨Ø§Øª Ù…Ø¶Ø§ÙØ©."
    lines = ["ğŸ‘¥ Ø§Ù„Ù‚Ø±ÙˆØ¨Ø§Øª:", ""]
    for i, g in enumerate(groups, start=1):
        lines.append(f"{i}) {g.get('name','Ù‚Ø±ÙˆØ¨')}")
        if g.get("telegram"):
            lines.append(f"   ØªÙŠÙ„ÙŠØ¬Ø±Ø§Ù…: {g.get('telegram')}")
        if g.get("whatsapp"):
            lines.append(f"   ÙˆØ§ØªØ³Ø§Ø¨: {g.get('whatsapp')}")
        lines.append(f"   ID: {g.get('id')}")
        lines.append("")
    return clip("\n".join(lines))


async def show_groups_platform(update: Update, platform: str) -> None:
    if not update.message:
        return
    groups = data_read(lambda: list(DATA.general_groups))
    if not groups:
        await update.message.reply_text("âŒ Ù„Ø§ ØªÙˆØ¬Ø¯ Ù‚Ø±ÙˆØ¨Ø§Øª Ù…Ø¶Ø§ÙØ© Ø­Ø§Ù„ÙŠÙ‹Ø§.")
        return

    shown = 0
    for g in groups[:60]:
        t = g.get("telegram", "")
        w = g.get("whatsapp", "")
        if platform == "tg" and not t:
            continue
        if platform == "wa" and not w:
            continue

        txt = f"ğŸ‘¥ {g.get('name','Ù‚Ø±ÙˆØ¨')}\n"
        if t:
            txt += f"ØªÙŠÙ„ÙŠØ¬Ø±Ø§Ù…: {t}\n"
        if w:
            txt += f"ÙˆØ§ØªØ³Ø§Ø¨: {w}\n"
        await update.message.reply_text(clip(txt.strip()), reply_markup=open_buttons(t, w))
        shown += 1
        if shown >= 12:
            break

    if shown == 0:
        await update.message.reply_text("âŒ Ù„Ø§ ØªÙˆØ¬Ø¯ Ù‚Ø±ÙˆØ¨Ø§Øª Ø¹Ù„Ù‰ Ù‡Ø°Ù‡ Ø§Ù„Ù…Ù†ØµØ© Ø­Ø§Ù„ÙŠØ§Ù‹.")


# =========================================================
# HIDE / RENAME MENUS TEXT
# =========================================================
def hide_menu_text() -> str:
    hidden = set(data_read(lambda: list(DATA.hidden_buttons)))
    lines = ["ğŸ‘ï¸â€ğŸ—¨ï¸ Ø¥Ø®ÙØ§Ø¡/Ø¥Ø¸Ù‡Ø§Ø± Ø§Ù„Ø£Ø²Ø±Ø§Ø± Ù„Ù„Ø¹Ø§Ù…Ø©", "", "Ø£Ø±Ø³Ù„ Ø±Ù‚Ù… Ø§Ù„Ø²Ø± Ù„ØªØ¨Ø¯ÙŠÙ„ Ø­Ø§Ù„ØªÙ‡:", ""]
    for i, k in enumerate(HIDEABLE_KEYS, start=1):
        status = "ğŸš« Ù…Ø®ÙÙŠ" if k in hidden else "âœ… Ø¸Ø§Ù‡Ø±"
        lines.append(f"{i}) {label_for(k)} â€” {status}")
    return clip("\n".join(lines))


def rename_menu_text() -> str:
    lines = ["âœï¸ ØªØ¹Ø¯ÙŠÙ„ Ø£Ø³Ù…Ø§Ø¡ Ø§Ù„Ø£Ø²Ø±Ø§Ø±", "", "Ø£Ø±Ø³Ù„ Ø±Ù‚Ù… Ø§Ù„Ø²Ø± Ù„ØªØ¹Ø¯ÙŠÙ„ Ø§Ø³Ù…Ù‡:", ""]
    for i, k in enumerate(HIDEABLE_KEYS, start=1):
        lines.append(f"{i}) {label_for(k)}")
    lines.append("")
    lines.append("Ù…Ù„Ø§Ø­Ø¸Ø©: Ø¥Ø±Ø³Ø§Ù„ Ø§Ø³Ù… ÙØ§Ø±Øº ÙŠØ¹ÙŠØ¯ Ø§Ù„Ø§Ø³Ù… Ø§Ù„Ø§ÙØªØ±Ø§Ø¶ÙŠ.")
    return clip("\n".join(lines))


# =========================================================
# BACK NAV
# =========================================================
async def go_back(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not update.message or not update.effective_user:
        return
    uid = update.effective_user.id
    mode = get_mode(context)

    if contact_gate_required(update):
        await update.message.reply_text("ğŸ”’ ÙŠÙ„Ø²Ù… Ù…Ø´Ø§Ø±ÙƒØ© Ø±Ù‚Ù…Ùƒ Ø£ÙˆÙ„Ù‹Ø§.", reply_markup=contact_kb())
        return

    if mode == MODE_CONFIRM:
        await confirm_return(update, context)
        return

    selected_college = get_selected_college(context)
    if selected_college:
        set_selected_college(context, None)
        set_mode(context, MODE_COLLEGES)
        await update.message.reply_text("ğŸ« Ø§Ø®ØªØ± Ø§Ù„ÙƒÙ„ÙŠØ©:", reply_markup=kb_colleges(uid))
        return

    if mode in {MODE_SERVICES, MODE_SUMMARIES, MODE_CONTACT_MENU, MODE_GG_USER_MENU, MODE_COLLEGES}:
        set_mode(context, MODE_NORMAL)
        set_selected_content(context, None, None)
        await update.message.reply_text("ğŸ ", reply_markup=kb_main(uid))
        return

    if mode in {MODE_CONTENT_EDIT_TEXT, MODE_CONTENT_UPLOAD_FILE}:
        parent = get_parent_menu_text(context) or label_for(KEY_MAIN_SERVICES)
        set_mode(context, MODE_NORMAL)
        # Ø§Ù„Ø¹ÙˆØ¯Ø© Ø­Ø³Ø¨ Ø§Ù„Ø£Ø¨
        if parent == label_for(KEY_MAIN_SUMMARIES):
            set_mode(context, MODE_SUMMARIES)
            await update.message.reply_text("ğŸ—‚ï¸", reply_markup=kb_summaries(uid))
        elif parent == label_for(KEY_MAIN_SERVICES):
            set_mode(context, MODE_SERVICES)
            await update.message.reply_text("ğŸ§°", reply_markup=kb_services(uid))
        else:
            # Ù…Ø«Ø§Ù„: Ø§Ù„ØªÙ‚ÙˆÙŠÙ… Ø£Ùˆ Ø§Ù„Ø³Ø­Ø¨ ÙˆØ§Ù„Ø¥Ø¶Ø§ÙØ©
            await update.message.reply_text("ğŸ ", reply_markup=kb_main(uid))
        return

    if mode in {MODE_SUPPORT, MODE_CC_MENU, MODE_GG_MENU, MODE_HIDE_MENU, MODE_RENAME_MENU, MODE_RENAME_VALUE, MODE_CAL_MANUAL_WAIT, MODE_AM_MENU, MODE_AM_ADD, MODE_AM_DEL}:
        set_mode(context, MODE_NORMAL)
        if is_admin(uid):
            await update.message.reply_text("âš™ï¸", reply_markup=kb_admin(uid))
        else:
            await update.message.reply_text("ğŸ ", reply_markup=kb_main(uid))
        return

    set_mode(context, MODE_NORMAL)
    await update.message.reply_text("ğŸ ", reply_markup=kb_main(uid))


# =========================================================
# TEXT ROUTER
# =========================================================
async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not update.message or not update.effective_user:
        return

    user = update.effective_user
    uid = user.id
    text = (update.message.text or "").strip()
    data_mutate(lambda: DATA.users.add(uid))

    # HOME
    if text == BTN_HOME:
        reset_flow(context)
        if contact_gate_required(update):
            await update.message.reply_text(WELCOME_TEXT)
            await update.message.reply_text(CONTACT_PROMPT_TEXT)
            await update.message.reply_text(IMPORTANT_NOTICE_TEXT, reply_markup=contact_kb())
            return
        await update.message.reply_text(WELCOME_TEXT, reply_markup=kb_main(uid))
        return

    # BACK
    if text == BTN_BACK:
        await go_back(update, context)
        return

    # BLOCK if gate required
    if contact_gate_required(update):
        await update.message.reply_text("ğŸ”’ ÙŠÙ„Ø²Ù… Ù…Ø´Ø§Ø±ÙƒØ© Ø±Ù‚Ù…Ùƒ Ø£ÙˆÙ„Ù‹Ø§ Ù„Ù„Ù…ØªØ§Ø¨Ø¹Ø©.", reply_markup=contact_kb())
        return

    mode = get_mode(context)

    # CANCEL
    if text == BTN_CANCEL and mode != MODE_NORMAL:
        set_mode(context, MODE_NORMAL)
        context.user_data.pop(USER_STEP, None)
        context.user_data.pop(USER_TEMP, None)
        context.user_data.pop(USER_SELECTED_ID, None)
        context.user_data.pop(USER_CHOICE, None)
        context.user_data.pop(USER_RENAME_KEY, None)
        await update.message.reply_text("âœ… ØªÙ… Ø§Ù„Ø¥Ù„ØºØ§Ø¡.", reply_markup=kb_main(uid))
        return

    # CONFIRM
    if mode == MODE_CONFIRM:
        if text == BTN_YES:
            await run_confirm_action(update, context)
            return
        if text == BTN_NO:
            await update.message.reply_text("âœ… ØªÙ… Ø§Ù„Ø¥Ù„ØºØ§Ø¡.", reply_markup=kb_main(uid))
            await confirm_return(update, context)
            return
        await update.message.reply_text("Ø§Ø®ØªØ± âœ… Ù†Ø¹Ù… Ø£Ùˆ âŒ Ù„Ø§.", reply_markup=kb_confirm())
        return

    # Block links for non-admin (Ù„ÙƒÙ† Ø§Ù„Ø±ÙˆØ§Ø¨Ø· Ù…Ù…ÙƒÙ† ØªØ¬ÙŠ Ù…Ù† Ø§Ù„Ø¨ÙˆØª Ù†ÙØ³Ù‡ Ø¹Ø¨Ø± Ø§Ù„Ù…Ø­ØªÙˆÙ‰)
    if not is_admin(uid) and contains_link(text):
        await update.message.reply_text("ğŸ”’ Ù…Ù…Ù†ÙˆØ¹ Ø¥Ø±Ø³Ø§Ù„ Ø±ÙˆØ§Ø¨Ø· Ø¯Ø§Ø®Ù„ Ø§Ù„Ø¨ÙˆØª.", reply_markup=kb_main(uid))
        return

    # SUPPORT MODE
    if mode == MODE_SUPPORT:
        item = {
            "time": now_str(),
            "name": user.full_name or "â€”",
            "username": f"@{user.username}" if user.username else "Ø¨Ø¯ÙˆÙ† Ù…Ø¹Ø±Ù",
            "id": str(uid),
            "text": clip(text),
        }
        push_inbox(item)
        msg = (
            "ğŸ“© Ø±Ø³Ø§Ù„Ø© Ø¬Ø¯ÙŠØ¯Ø©\n\n"
            f"Ø§Ù„Ø§Ø³Ù…: {item['name']}\n"
            f"Ø§Ù„Ù…Ø¹Ø±Ù: {item['username']}\n"
            f"ID: {uid}\n\n"
            f"Ø§Ù„Ù†Øµ:\n{text}"
        )
        await forward_to_admins(context.application, msg)
        set_mode(context, MODE_NORMAL)
        await update.message.reply_text("âœ… ØªÙ… Ø¥Ø±Ø³Ø§Ù„ Ø±Ø³Ø§Ù„ØªÙƒ Ù„Ù„Ø¥Ø¯Ø§Ø±Ø©.", reply_markup=kb_main(uid))
        return

    # HIDE MENU (super only)
    if mode == MODE_HIDE_MENU:
        if not is_super_admin(uid):
            await update.message.reply_text(deny_no_perm(), reply_markup=kb_admin(uid))
            set_mode(context, MODE_NORMAL)
            return
        idx = safe_int(text)
        if not idx or idx < 1 or idx > len(HIDEABLE_KEYS):
            await update.message.reply_text("Ø£Ø±Ø³Ù„ Ø±Ù‚Ù… ØµØ­ÙŠØ­.", reply_markup=kb_wait_cancel())
            return
        key = HIDEABLE_KEYS[idx - 1]
        start_confirm(context, "toggle_hide", {"key": key}, {"target": "admin"})
        await update.message.reply_text(f"ØªØ£ÙƒÙŠØ¯ ØªØ¨Ø¯ÙŠÙ„ Ø­Ø§Ù„Ø©: {label_for(key)} ØŸ", reply_markup=kb_confirm())
        return

    # RENAME MENU (super only)
    if mode == MODE_RENAME_MENU:
        if not is_super_admin(uid):
            await update.message.reply_text(deny_no_perm(), reply_markup=kb_admin(uid))
            set_mode(context, MODE_NORMAL)
            return
        idx = safe_int(text)
        if not idx or idx < 1 or idx > len(HIDEABLE_KEYS):
            await update.message.reply_text("Ø£Ø±Ø³Ù„ Ø±Ù‚Ù… ØµØ­ÙŠØ­.", reply_markup=kb_wait_cancel())
            return
        key = HIDEABLE_KEYS[idx - 1]
        context.user_data[USER_RENAME_KEY] = key
        set_mode(context, MODE_RENAME_VALUE)
        await update.message.reply_text(f"Ø£Ø±Ø³Ù„ Ø§Ù„Ø§Ø³Ù… Ø§Ù„Ø¬Ø¯ÙŠØ¯ Ù„Ù„Ø²Ø±:\n{label_for(key)}", reply_markup=kb_wait_cancel())
        return

    if mode == MODE_RENAME_VALUE:
        if not is_super_admin(uid):
            await update.message.reply_text(deny_no_perm(), reply_markup=kb_admin(uid))
            set_mode(context, MODE_NORMAL)
            return
        key = str(context.user_data.get(USER_RENAME_KEY, "")).strip()
        if key not in HIDEABLE_KEYS:
            set_mode(context, MODE_NORMAL)
            await update.message.reply_text("âŒ Ø®Ø·Ø£.", reply_markup=kb_admin(uid))
            return
        start_confirm(context, "rename_button", {"key": key, "value": text.strip()}, {"target": "admin"})
        await update.message.reply_text("ØªØ£ÙƒÙŠØ¯ ØªØ¹Ø¯ÙŠÙ„ Ø§Ø³Ù… Ø§Ù„Ø²Ø±ØŸ", reply_markup=kb_confirm())
        return

    # CALENDAR MANUAL WAIT (super only) - legacy manual text
    if mode == MODE_CAL_MANUAL_WAIT:
        if not is_super_admin(uid):
            await update.message.reply_text(deny_no_perm(), reply_markup=kb_admin(uid))
            set_mode(context, MODE_NORMAL)
            return
        manual = text.strip()

        def _mut():
            DATA.calendar["display_mode"] = "manual"
            DATA.calendar["manual_text"] = manual
            DATA.calendar["last_updated"] = now_str()
            DATA.calendar["last_source"] = "manual"

        data_mutate(_mut)
        set_mode(context, MODE_NORMAL)
        await update.message.reply_text("âœ… ØªÙ… Ø­ÙØ¸ Ø§Ù„ØªÙ‚ÙˆÙŠÙ… Ø§Ù„Ù…Ø®ØµØµ.", reply_markup=kb_admin(uid))
        return

    # CONTENT EDIT TEXT
    if mode == MODE_CONTENT_EDIT_TEXT:
        if not is_admin(uid):
            await update.message.reply_text("âŒ ØºÙŠØ± Ù…ØªØ§Ø­.", reply_markup=kb_main(uid))
            set_mode(context, MODE_NORMAL)
            return
        item_key = get_selected_itemkey(context)
        parent_menu = get_parent_menu_text(context) or label_for(KEY_MAIN_SERVICES)
        if not item_key:
            set_mode(context, MODE_NORMAL)
            await update.message.reply_text("âŒ Ø®Ø·Ø£.", reply_markup=kb_main(uid))
            return
        set_content_text(item_key, text)
        set_mode(context, MODE_NORMAL)
        await update.message.reply_text("âœ… ØªÙ… ØªØ¹Ø¯ÙŠÙ„ Ø§Ù„Ù†Øµ.", reply_markup=kb_content_admin(parent_menu))
        return

    # COLLEGES ADD/DEL (admin)
    selected_college = get_selected_college(context)
    if mode in {MODE_ADD_WA, MODE_DEL_WA, MODE_ADD_TG, MODE_DEL_TG}:
        if not is_admin(uid) or not selected_college:
            set_mode(context, MODE_NORMAL)
            await update.message.reply_text("âŒ ØºÙŠØ± Ù…ØªØ§Ø­.", reply_markup=kb_main(uid))
            return

        value = text.strip()

        if mode == MODE_ADD_WA:
            if not normalize_whatsapp_link(value):
                await update.message.reply_text("âŒ Ø±Ø§Ø¨Ø· ÙˆØ§ØªØ³Ø§Ø¨ ØºÙŠØ± ØµØ­ÙŠØ­.", reply_markup=kb_wait_cancel())
                return
            start_confirm(context, "college_add_wa", {"cname": selected_college, "link": value}, {"target": "college_view"})
            await update.message.reply_text("ØªØ£ÙƒÙŠØ¯ Ø¥Ø¶Ø§ÙØ© Ø±Ø§Ø¨Ø· ÙˆØ§ØªØ³Ø§Ø¨ØŸ", reply_markup=kb_confirm())
            return

        if mode == MODE_DEL_WA:
            start_confirm(context, "college_del_wa", {"cname": selected_college, "link": value}, {"target": "college_view"})
            await update.message.reply_text("ØªØ£ÙƒÙŠØ¯ Ø­Ø°Ù Ø±Ø§Ø¨Ø· ÙˆØ§ØªØ³Ø§Ø¨ØŸ", reply_markup=kb_confirm())
            return

        if mode == MODE_ADD_TG:
            if not normalize_telegram_link(value):
                await update.message.reply_text("âŒ Ø±Ø§Ø¨Ø·/Ù…Ø¹Ø±Ù ØªÙŠÙ„ÙŠØ¬Ø±Ø§Ù… ØºÙŠØ± ØµØ­ÙŠØ­.", reply_markup=kb_wait_cancel())
                return
            start_confirm(context, "college_add_tg", {"cname": selected_college, "link": value}, {"target": "college_view"})
            await update.message.reply_text("ØªØ£ÙƒÙŠØ¯ Ø¥Ø¶Ø§ÙØ© Ø±Ø§Ø¨Ø· ØªÙŠÙ„ÙŠØ¬Ø±Ø§Ù…ØŸ", reply_markup=kb_confirm())
            return

        if mode == MODE_DEL_TG:
            start_confirm(context, "college_del_tg", {"cname": selected_college, "link": value}, {"target": "college_view"})
            await update.message.reply_text("ØªØ£ÙƒÙŠØ¯ Ø­Ø°Ù Ø±Ø§Ø¨Ø· ØªÙŠÙ„ÙŠØ¬Ø±Ø§Ù…ØŸ", reply_markup=kb_confirm())
            return

    # GENERAL GROUPS ADMIN FLOWS
    if mode == MODE_GG_ADD and is_admin(uid):
        step = context.user_data.get(USER_STEP)
        temp: dict = context.user_data.get(USER_TEMP, {})

        if step == "name":
            temp["name"] = text.strip() or "Ù‚Ø±ÙˆØ¨"
            context.user_data[USER_STEP] = "telegram"
            context.user_data[USER_TEMP] = temp
            await update.message.reply_text("Ø£Ø±Ø³Ù„ Ø±Ø§Ø¨Ø· ØªÙŠÙ„ÙŠØ¬Ø±Ø§Ù… Ù„Ù„Ù‚Ø±ÙˆØ¨ (t.me Ø£Ùˆ @username):", reply_markup=kb_wait_cancel())
            return

        if step == "telegram":
            t = normalize_telegram_link(text)
            if not t:
                await update.message.reply_text("âŒ Ø±Ø§Ø¨Ø· ØªÙŠÙ„ÙŠØ¬Ø±Ø§Ù… ØºÙŠØ± ØµØ­ÙŠØ­.", reply_markup=kb_wait_cancel())
                return
            temp["telegram"] = t
            context.user_data[USER_STEP] = "whatsapp"
            context.user_data[USER_TEMP] = temp
            await update.message.reply_text("Ø£Ø±Ø³Ù„ Ø±Ø§Ø¨Ø· ÙˆØ§ØªØ³Ø§Ø¨ (Ø§Ø®ØªÙŠØ§Ø±ÙŠ). Ø§ÙƒØªØ¨ - Ù„ØªØ®Ø·ÙŠ:", reply_markup=kb_wait_cancel())
            return

        if step == "whatsapp":
            w_raw = text.strip()
            w = "" if w_raw == "-" else (normalize_whatsapp_link(w_raw) or None)
            if w_raw != "-" and w is None:
                await update.message.reply_text("âŒ Ø±Ø§Ø¨Ø· ÙˆØ§ØªØ³Ø§Ø¨ ØºÙŠØ± ØµØ­ÙŠØ­. Ø£Ùˆ Ø§ÙƒØªØ¨ - Ù„Ù„ØªØ®Ø·ÙŠ", reply_markup=kb_wait_cancel())
                return

            group = {
                "id": str(int(time.time() * 1000)),
                "name": str(temp.get("name", "Ù‚Ø±ÙˆØ¨")),
                "telegram": str(temp.get("telegram", "")),
                "whatsapp": w or "",
            }
            context.user_data.pop(USER_STEP, None)
            context.user_data.pop(USER_TEMP, None)
            start_confirm(context, "gg_add", {"group": group}, {"target": "gg_admin"})
            await update.message.reply_text("ØªØ£ÙƒÙŠØ¯ Ø¥Ø¶Ø§ÙØ© Ø§Ù„Ù‚Ø±ÙˆØ¨ØŸ", reply_markup=kb_confirm())
            return

    if mode == MODE_GG_DEL and is_admin(uid):
        gg = _gg_find(text.strip())
        if not gg:
            await update.message.reply_text("âŒ ID ØºÙŠØ± ØµØ­ÙŠØ­.", reply_markup=kb_wait_cancel())
            return
        start_confirm(context, "gg_del", {"id": gg.get("id")}, {"target": "gg_admin"})
        await update.message.reply_text("ØªØ£ÙƒÙŠØ¯ Ø­Ø°Ù Ø§Ù„Ù‚Ø±ÙˆØ¨ØŸ", reply_markup=kb_confirm())
        return

    if mode == MODE_GG_EDIT_SELECT and is_admin(uid):
        gg = _gg_find(text.strip())
        if not gg:
            await update.message.reply_text("âŒ ID ØºÙŠØ± ØµØ­ÙŠØ­.", reply_markup=kb_wait_cancel())
            return
        context.user_data[USER_SELECTED_ID] = gg["id"]
        set_mode(context, MODE_GG_EDIT_FIELD)
        await update.message.reply_text("Ø§ÙƒØªØ¨ Ø±Ù‚Ù… Ø§Ù„Ø®ÙŠØ§Ø±:\n1) Ø§Ù„Ø§Ø³Ù…\n2) ØªÙŠÙ„ÙŠØ¬Ø±Ø§Ù…\n3) ÙˆØ§ØªØ³Ø§Ø¨", reply_markup=kb_wait_cancel())
        return

    if mode == MODE_GG_EDIT_FIELD and is_admin(uid):
        choice = text.strip()
        if choice not in {"1", "2", "3"}:
            await update.message.reply_text("Ø§ÙƒØªØ¨ 1 Ø£Ùˆ 2 Ø£Ùˆ 3.", reply_markup=kb_wait_cancel())
            return
        context.user_data[USER_CHOICE] = choice
        set_mode(context, MODE_GG_EDIT_VALUE)
        if choice == "1":
            await update.message.reply_text("Ø£Ø±Ø³Ù„ Ø§Ù„Ø§Ø³Ù… Ø§Ù„Ø¬Ø¯ÙŠØ¯:", reply_markup=kb_wait_cancel())
        elif choice == "2":
            await update.message.reply_text("Ø£Ø±Ø³Ù„ Ø±Ø§Ø¨Ø· ØªÙŠÙ„ÙŠØ¬Ø±Ø§Ù… Ø§Ù„Ø¬Ø¯ÙŠØ¯:", reply_markup=kb_wait_cancel())
        else:
            await update.message.reply_text("Ø£Ø±Ø³Ù„ Ø±Ø§Ø¨Ø· ÙˆØ§ØªØ³Ø§Ø¨ Ø§Ù„Ø¬Ø¯ÙŠØ¯ (Ø£Ùˆ - Ù„Ù…Ø³Ø­Ù‡):", reply_markup=kb_wait_cancel())
        return

    if mode == MODE_GG_EDIT_VALUE and is_admin(uid):
        gg_id = str(context.user_data.get(USER_SELECTED_ID, ""))
        gg = _gg_find(gg_id)
        if not gg:
            set_mode(context, MODE_GG_MENU)
            await update.message.reply_text("âŒ Ù„Ù… Ø£Ø¬Ø¯ Ø§Ù„Ù‚Ø±ÙˆØ¨.", reply_markup=kb_general_groups_admin())
            return

        choice = context.user_data.get(USER_CHOICE)
        if choice == "1":
            gg["name"] = text.strip() or gg.get("name", "Ù‚Ø±ÙˆØ¨")
        elif choice == "2":
            t = normalize_telegram_link(text)
            if not t:
                await update.message.reply_text("âŒ Ø±Ø§Ø¨Ø· ØªÙŠÙ„ÙŠØ¬Ø±Ø§Ù… ØºÙŠØ± ØµØ­ÙŠØ­.", reply_markup=kb_wait_cancel())
                return
            gg["telegram"] = t
        else:
            if text.strip() == "-":
                gg["whatsapp"] = ""
            else:
                w = normalize_whatsapp_link(text)
                if not w:
                    await update.message.reply_text("âŒ Ø±Ø§Ø¨Ø· ÙˆØ§ØªØ³Ø§Ø¨ ØºÙŠØ± ØµØ­ÙŠØ­. Ø£Ùˆ Ø§ÙƒØªØ¨ - Ù„Ù…Ø³Ø­Ù‡.", reply_markup=kb_wait_cancel())
                    return
                gg["whatsapp"] = w

        def _mut():
            for i, x in enumerate(DATA.general_groups):
                if str(x.get("id")) == gg_id:
                    DATA.general_groups[i] = gg
                    break

        data_mutate(_mut)
        set_mode(context, MODE_GG_MENU)
        await update.message.reply_text("âœ… ØªÙ… Ø§Ù„ØªØ¹Ø¯ÙŠÙ„.", reply_markup=kb_general_groups_admin())
        return

    # ADMIN MANAGE (super)
    if mode == MODE_AM_ADD:
        if not is_super_admin(uid):
            await update.message.reply_text(deny_no_perm(), reply_markup=kb_admin(uid))
            set_mode(context, MODE_NORMAL)
            return
        new_id = safe_int(text)
        if not new_id:
            await update.message.reply_text("Ø£Ø±Ø³Ù„ Ø±Ù‚Ù… ID ØµØ­ÙŠØ­.", reply_markup=kb_wait_cancel())
            return
        start_confirm(context, "admin_add", {"id": str(new_id)}, {"target": "am_menu"})
        await update.message.reply_text("ØªØ£ÙƒÙŠØ¯ Ø¥Ø¶Ø§ÙØ© Ø§Ù„Ø£Ø¯Ù…Ù†ØŸ", reply_markup=kb_confirm())
        return

    if mode == MODE_AM_DEL:
        if not is_super_admin(uid):
            await update.message.reply_text(deny_no_perm(), reply_markup=kb_admin(uid))
            set_mode(context, MODE_NORMAL)
            return
        rem_id = safe_int(text)
        if not rem_id:
            await update.message.reply_text("Ø£Ø±Ø³Ù„ Ø±Ù‚Ù… ID ØµØ­ÙŠØ­.", reply_markup=kb_wait_cancel())
            return
        start_confirm(context, "admin_del", {"id": str(rem_id)}, {"target": "am_menu"})
        await update.message.reply_text("ØªØ£ÙƒÙŠØ¯ Ø­Ø°Ù Ø§Ù„Ø£Ø¯Ù…Ù†ØŸ", reply_markup=kb_confirm())
        return

    # CONTENT ADMIN PANEL BUTTONS (works for: services/summaries/calendar/add-drop)
    item_key = get_selected_itemkey(context)
    parent_menu_text = get_parent_menu_text(context) or label_for(KEY_MAIN_SERVICES)
    if is_admin(uid) and item_key and text in {BTN_UPLOAD_FILE, BTN_EDIT_TEXT, BTN_DEL_FILE, BTN_DELETE_SECTION}:
        if text == BTN_UPLOAD_FILE:
            set_mode(context, MODE_CONTENT_UPLOAD_FILE)
            await update.message.reply_text("ğŸ“ Ø£Ø±Ø³Ù„ Ù…Ù„Ù Ø£Ùˆ ØµÙˆØ±Ø© Ø§Ù„Ø¢Ù†.\nÙ„Ù„Ø¥Ù„ØºØ§Ø¡: âŒ Ø¥Ù„ØºØ§Ø¡", reply_markup=kb_wait_cancel())
            return
        if text == BTN_EDIT_TEXT:
            set_mode(context, MODE_CONTENT_EDIT_TEXT)
            await update.message.reply_text("âœï¸ Ø£Ø±Ø³Ù„ Ø§Ù„Ù†Øµ Ø§Ù„Ø¬Ø¯ÙŠØ¯:", reply_markup=kb_wait_cancel())
            return
        if text == BTN_DEL_FILE:
            start_confirm(context, "content_del_file", {"item_key": item_key}, {"target": "main"})
            await update.message.reply_text("ØªØ£ÙƒÙŠØ¯ Ø­Ø°Ù Ø§Ù„Ù…Ù„ÙØŸ", reply_markup=kb_confirm())
            return
        if text == BTN_DELETE_SECTION:
            start_confirm(context, "content_del_section", {"item_key": item_key}, {"target": "main"})
            await update.message.reply_text("ØªØ£ÙƒÙŠØ¯ Ø­Ø°Ù Ø§Ù„Ù‚Ø³Ù… Ø¨Ø§Ù„ÙƒØ§Ù…Ù„ØŸ", reply_markup=kb_confirm())
            return

    # MAIN MENU (dynamic)
    k_main = resolve_key_by_text(
        text,
        [
            KEY_MAIN_COLLEGES,
            KEY_MAIN_ADMISSION,
            KEY_MAIN_MAJORS,
            KEY_MAIN_ADD_DROP,
            KEY_MAIN_CALENDAR,
            KEY_MAIN_SUMMARIES,
            KEY_MAIN_SERVICES,
            KEY_MAIN_GROUPS,
            KEY_MAIN_CONTACT,
            KEY_MAIN_ABOUT,
        ],
    )

    if k_main == KEY_MAIN_ADMISSION:
        await update.message.reply_text(AOU_ADMISSION, reply_markup=kb_main(uid))
        return

    if k_main == KEY_MAIN_MAJORS:
        await update.message.reply_text(AOU_MAJORS, reply_markup=kb_main(uid))
        return

    if k_main == KEY_MAIN_ABOUT:
        await update.message.reply_text(AOU_ABOUT, reply_markup=kb_main(uid))
        return

    # âœ… Ø§Ù„Ø³Ø­Ø¨ ÙˆØ§Ù„Ø¥Ø¶Ø§ÙØ©: Ù…Ø­ØªÙˆÙ‰ Ù‚Ø§Ø¨Ù„ Ù„Ù„Ø¥Ø¯Ø§Ø±Ø© Ù„Ù„Ø£Ø¯Ù…Ù† ÙÙ‚Ø·
    if k_main == KEY_MAIN_ADD_DROP:
        set_selected_content(context, CONTENT_KEY_ADD_DROP, label_for(KEY_MAIN_ADD_DROP))
        await send_content_to_user(update, CONTENT_KEY_ADD_DROP, fallback_text=AOU_ADD_DROP_FALLBACK)
        if is_admin(uid):
            await update.message.reply_text("Ù„ÙˆØ­Ø© Ø¥Ø¯Ø§Ø±Ø© Ø§Ù„Ø³Ø­Ø¨ ÙˆØ§Ù„Ø¥Ø¶Ø§ÙØ©:", reply_markup=kb_content_admin(label_for(KEY_MAIN_ADD_DROP)))
        else:
            await update.message.reply_text("â¬…ï¸", reply_markup=kb_main(uid))
        return

    # âœ… Ø§Ù„ØªÙ‚ÙˆÙŠÙ…: ÙŠØ¹Ø±Ø¶ Ù…Ø±ÙÙ‚/Ù†Øµ Ø§Ù„Ø£Ø¯Ù…Ù† Ø¥Ù† ÙˆØ¬Ø¯ØŒ ÙˆØ¥Ù„Ø§ ÙŠØ¹Ø±Ø¶ Ø§Ù„Ø±ÙˆØ§Ø¨Ø· Ø§Ù„ØªÙ„Ù‚Ø§Ø¦ÙŠØ©
    if k_main == KEY_MAIN_CALENDAR:
        set_selected_content(context, CONTENT_KEY_CALENDAR_ATTACH, label_for(KEY_MAIN_CALENDAR))
        if not content_is_empty(CONTENT_KEY_CALENDAR_ATTACH):
            await send_content_to_user(update, CONTENT_KEY_CALENDAR_ATTACH, fallback_text=None)
        else:
            await update.message.reply_text(calendar_text_plain(), reply_markup=kb_main(uid), disable_web_page_preview=True)
            kb = calendar_links_buttons()
            if kb:
                await update.message.reply_text("â¬‡ï¸ ÙØªØ­ Ù…Ù„ÙØ§Øª Ø§Ù„ØªÙ‚ÙˆÙŠÙ…:", reply_markup=kb)

        if is_admin(uid):
            await update.message.reply_text("Ù„ÙˆØ­Ø© Ø¥Ø¯Ø§Ø±Ø© Ø§Ù„ØªÙ‚ÙˆÙŠÙ…:", reply_markup=kb_content_admin(label_for(KEY_MAIN_CALENDAR)))
        else:
            await update.message.reply_text("â¬…ï¸", reply_markup=kb_main(uid))
        return

    if k_main == KEY_MAIN_CONTACT:
        set_mode(context, MODE_CONTACT_MENU)
        await update.message.reply_text("ğŸ“ Ø§Ø®ØªØ±:", reply_markup=kb_contact(uid))
        return

    if k_main == KEY_MAIN_SERVICES:
        set_mode(context, MODE_SERVICES)
        set_selected_content(context, None, None)
        await update.message.reply_text("ğŸ§° Ø§Ø®ØªØ±:", reply_markup=kb_services(uid))
        return

    if k_main == KEY_MAIN_SUMMARIES:
        set_mode(context, MODE_SUMMARIES)
        set_selected_content(context, None, None)
        await update.message.reply_text("ğŸ—‚ï¸ Ø§Ø®ØªØ±:", reply_markup=kb_summaries(uid))
        return

    if k_main == KEY_MAIN_GROUPS:
        set_mode(context, MODE_GG_USER_MENU)
        await update.message.reply_text("ğŸ‘¥ Ø§Ø®ØªØ± Ø§Ù„Ù…Ù†ØµØ©:", reply_markup=kb_general_groups_user(uid))
        return

    if k_main == KEY_MAIN_COLLEGES:
        set_mode(context, MODE_COLLEGES)
        set_selected_college(context, None)
        await update.message.reply_text("ğŸ« Ø§Ø®ØªØ± Ø§Ù„ÙƒÙ„ÙŠØ©:", reply_markup=kb_colleges(uid))
        return

    # CONTACT MENU
    if text == BTN_UNIV_NUMBERS:
        await update.message.reply_text(CONTACT_INFO, reply_markup=kb_contact(uid), disable_web_page_preview=True)
        return

    if text == BTN_SOCIALS:
        kb = InlineKeyboardMarkup(
            [
                [InlineKeyboardButton("ğŸ”— Linktree Ù…Ù†ØµØ§Øª Ø§Ù„Ø¬Ø§Ù…Ø¹Ø©", url=AOU_SOCIAL_LINKTREE)],
                [InlineKeyboardButton("ğŸŒ Ù…ÙˆÙ‚Ø¹ Ø§Ù„Ø¬Ø§Ù…Ø¹Ø©", url=UNIV_URL)],
            ]
        )
        await update.message.reply_text("ğŸ“² Ù…Ù†ØµØ§Øª Ø§Ù„Ø¬Ø§Ù…Ø¹Ø©:", reply_markup=kb)
        return

    if text == BTN_CONTACT_ADMIN:
        set_mode(context, MODE_SUPPORT)
        await update.message.reply_text("âœï¸ Ø§ÙƒØªØ¨ Ø±Ø³Ø§Ù„ØªÙƒ Ø§Ù„Ø¢Ù†.\nÙ„Ù„Ø¥Ù„ØºØ§Ø¡: âŒ Ø¥Ù„ØºØ§Ø¡", reply_markup=kb_wait_cancel())
        return

    # SERVICES ITEM SELECT
    k_serv = resolve_key_by_text(
        text,
        [
            KEY_SERV_SCHEDULE,
            KEY_SERV_ANNUAL,
            KEY_SERV_EXAMS,
            KEY_SERV_REG_CONT,
            KEY_SERV_REG_NEW,
            KEY_SERV_ABSENCE,
            KEY_SERV_DEPRIVATION,
            KEY_SERV_STRIKE,
        ],
    )
    if k_serv:
        item = CONTENT_KEYS[k_serv]
        set_selected_content(context, item, label_for(KEY_MAIN_SERVICES))
        await send_content_to_user(update, item)
        if is_admin(uid):
            await update.message.reply_text("Ù„ÙˆØ­Ø© Ø¥Ø¯Ø§Ø±Ø©:", reply_markup=kb_content_admin(label_for(KEY_MAIN_SERVICES)))
        else:
            set_mode(context, MODE_SERVICES)
            await update.message.reply_text("â¬…ï¸", reply_markup=kb_services(uid))
        return

    # SUMMARIES ITEM SELECT
    k_sum = resolve_key_by_text(text, [KEY_SUM_BOOKS, KEY_SUM_NOTES])
    if k_sum:
        item = CONTENT_KEYS[k_sum]
        set_selected_content(context, item, label_for(KEY_MAIN_SUMMARIES))
        await send_content_to_user(update, item)
        if is_admin(uid):
            await update.message.reply_text("Ù„ÙˆØ­Ø© Ø¥Ø¯Ø§Ø±Ø©:", reply_markup=kb_content_admin(label_for(KEY_MAIN_SUMMARIES)))
        else:
            set_mode(context, MODE_SUMMARIES)
            await update.message.reply_text("â¬…ï¸", reply_markup=kb_summaries(uid))
        return

    # GENERAL GROUPS USER MENU
    if mode == MODE_GG_USER_MENU and text in {BTN_GG_USER_TG, BTN_GG_USER_WA}:
        await show_groups_platform(update, "tg" if text == BTN_GG_USER_TG else "wa")
        await update.message.reply_text("â¬…ï¸", reply_markup=kb_general_groups_user(uid))
        return

    # COLLEGES: pick college
    if mode == MODE_COLLEGES:
        cols = data_read(lambda: dict(DATA.colleges))
        if text in cols:
            set_selected_college(context, text)
            set_mode(context, MODE_COLLEGE_VIEW)
            col = cols.get(text, {})
            await update.message.reply_text(
                clip(f"{text}\n\nğŸ“Œ {col.get('about','')}\n\nğŸ”— {col.get('url','ØºÙŠØ± Ù…ØªÙˆÙØ±')}"),
                reply_markup=kb_college_view(uid),
            )
            return

    # COLLEGE VIEW
    if get_mode(context) == MODE_COLLEGE_VIEW and (cname := get_selected_college(context)):
        col = data_read(lambda: dict(DATA.colleges.get(cname, {})))
        if text == BTN_COLLEGE_ABOUT:
            await update.message.reply_text(f"ğŸ“Œ {cname}\n\n{col.get('about','')}", reply_markup=kb_college_view(uid))
            return
        if text == BTN_COLLEGE_URL:
            await update.message.reply_text(f"ğŸ”— {cname}\n{col.get('url','ØºÙŠØ± Ù…ØªÙˆÙØ±')}", reply_markup=kb_college_view(uid))
            return
        if text == BTN_COLLEGE_WA:
            gs = col.get("whatsapp", [])
            if not gs:
                await update.message.reply_text("ğŸ“± Ù„Ø§ ØªÙˆØ¬Ø¯ Ù‚Ø±ÙˆØ¨Ø§Øª ÙˆØ§ØªØ³Ø§Ø¨ Ù…Ø¶Ø§ÙØ©.", reply_markup=kb_college_view(uid))
            else:
                await update.message.reply_text("ğŸ“± Ù‚Ø±ÙˆØ¨Ø§Øª ÙˆØ§ØªØ³Ø§Ø¨:\n" + "\n".join([f"- {g}" for g in gs]), reply_markup=kb_college_view(uid))
            return
        if text == BTN_COLLEGE_TG:
            gs = col.get("telegram", [])
            if not gs:
                await update.message.reply_text("ğŸ“¢ Ù„Ø§ ØªÙˆØ¬Ø¯ Ù‚Ø±ÙˆØ¨Ø§Øª ØªÙŠÙ„ÙŠØ¬Ø±Ø§Ù… Ù…Ø¶Ø§ÙØ©.", reply_markup=kb_college_view(uid))
            else:
                await update.message.reply_text("ğŸ“¢ Ù‚Ø±ÙˆØ¨Ø§Øª ØªÙŠÙ„ÙŠØ¬Ø±Ø§Ù…:\n" + "\n".join([f"- {g}" for g in gs]), reply_markup=kb_college_view(uid))
            return

        if is_admin(uid) and text == BTN_ADD_WA:
            set_mode(context, MODE_ADD_WA)
            await update.message.reply_text("ğŸ“± Ø£Ø±Ø³Ù„ Ø±Ø§Ø¨Ø· ÙˆØ§ØªØ³Ø§Ø¨:", reply_markup=kb_wait_cancel())
            return
        if is_admin(uid) and text == BTN_DEL_WA:
            set_mode(context, MODE_DEL_WA)
            await update.message.reply_text("ğŸ“± Ø£Ø±Ø³Ù„ Ø±Ø§Ø¨Ø· ÙˆØ§ØªØ³Ø§Ø¨ Ø§Ù„Ù…Ø±Ø§Ø¯ Ø­Ø°ÙÙ‡:", reply_markup=kb_wait_cancel())
            return
        if is_admin(uid) and text == BTN_ADD_TG:
            set_mode(context, MODE_ADD_TG)
            await update.message.reply_text("ğŸ“¢ Ø£Ø±Ø³Ù„ Ø±Ø§Ø¨Ø·/Ù…Ø¹Ø±Ù ØªÙŠÙ„ÙŠØ¬Ø±Ø§Ù…:", reply_markup=kb_wait_cancel())
            return
        if is_admin(uid) and text == BTN_DEL_TG:
            set_mode(context, MODE_DEL_TG)
            await update.message.reply_text("ğŸ“¢ Ø£Ø±Ø³Ù„ Ø±Ø§Ø¨Ø·/Ù…Ø¹Ø±Ù ØªÙŠÙ„ÙŠØ¬Ø±Ø§Ù… Ø§Ù„Ù…Ø±Ø§Ø¯ Ø­Ø°ÙÙ‡:", reply_markup=kb_wait_cancel())
            return

    # ADMIN SETTINGS ENTRY
    if text == BTN_ADMIN_SETTINGS:
        if not is_admin(uid):
            await update.message.reply_text("âŒ ØºÙŠØ± Ù…ØªØ§Ø­.", reply_markup=kb_main(uid))
            return
        await update.message.reply_text("âš™ï¸ Ø¥Ø¹Ø¯Ø§Ø¯Ø§Øª:", reply_markup=kb_admin(uid))
        return

    # ADMIN: inbox
    if is_admin(uid) and text == BTN_INBOX_SHOW:
        await update.message.reply_text(format_inbox_plain(50), reply_markup=kb_admin(uid))
        return

    if is_admin(uid) and text == BTN_INBOX_CLEAR:
        start_confirm(context, "inbox_clear", {}, {"target": "admin"})
        await update.message.reply_text("ØªØ£ÙƒÙŠØ¯ Ù…Ø³Ø­ Ø§Ù„Ø±Ø³Ø§Ø¦Ù„ØŸ", reply_markup=kb_confirm())
        return

    # ADMIN: calendar (admin settings screen)
    if is_admin(uid) and text == BTN_CAL_SHOW:
        # ÙŠØ¹Ø±Ø¶ Ø§Ù„Ù…Ø±ÙÙ‚ Ø¥Ù† ÙˆØ¬Ø¯ ÙˆØ¥Ù„Ø§ ÙŠØ¹Ø±Ø¶ Ø§Ù„Ø±ÙˆØ§Ø¨Ø·
        if not content_is_empty(CONTENT_KEY_CALENDAR_ATTACH):
            set_selected_content(context, CONTENT_KEY_CALENDAR_ATTACH, label_for(KEY_MAIN_CALENDAR))
            await send_content_to_user(update, CONTENT_KEY_CALENDAR_ATTACH)
        else:
            await update.message.reply_text(calendar_text_plain(), reply_markup=kb_admin(uid), disable_web_page_preview=True)
            kb = calendar_links_buttons()
            if kb:
                await update.message.reply_text("â¬‡ï¸ ÙØªØ­ Ù…Ù„ÙØ§Øª Ø§Ù„ØªÙ‚ÙˆÙŠÙ…:", reply_markup=kb)
        return

    if is_admin(uid) and text == BTN_CAL_REFRESH:
        ok, note = await asyncio.to_thread(refresh_calendar_links, True)
        await update.message.reply_text(note, reply_markup=kb_admin(uid))
        return

    if text == BTN_CAL_SET_MANUAL:
        if not is_super_admin(uid):
            await update.message.reply_text(deny_no_perm(), reply_markup=kb_admin(uid))
            return
        set_mode(context, MODE_CAL_MANUAL_WAIT)
        await update.message.reply_text("âœï¸ Ø£Ø±Ø³Ù„ Ù†Øµ Ø§Ù„ØªÙ‚ÙˆÙŠÙ… Ø§Ù„Ù…Ø®ØµØµ Ø§Ù„Ø¢Ù†:", reply_markup=kb_wait_cancel())
        return

    if text == BTN_CAL_USE_AUTO:
        if not is_super_admin(uid):
            await update.message.reply_text(deny_no_perm(), reply_markup=kb_admin(uid))
            return
        start_confirm(context, "calendar_use_auto", {}, {"target": "admin"})
        await update.message.reply_text("ØªØ£ÙƒÙŠØ¯ ØªØ­ÙˆÙŠÙ„ Ø§Ù„ØªÙ‚ÙˆÙŠÙ… Ø¥Ù„Ù‰ ØªÙ„Ù‚Ø§Ø¦ÙŠØŸ", reply_markup=kb_confirm())
        return

    if text == BTN_CAL_CLEAR:
        if not is_super_admin(uid):
            await update.message.reply_text(deny_no_perm(), reply_markup=kb_admin(uid))
            return
        start_confirm(context, "calendar_clear", {}, {"target": "admin"})
        await update.message.reply_text("ØªØ£ÙƒÙŠØ¯ Ø­Ø°Ù/Ø¥Ø®ÙØ§Ø¡ Ø§Ù„ØªÙ‚ÙˆÙŠÙ…ØŸ", reply_markup=kb_confirm())
        return

    if text == BTN_CAL_AUTO_ON:
        if not is_super_admin(uid):
            await update.message.reply_text(deny_no_perm(), reply_markup=kb_admin(uid))
            return
        start_confirm(context, "calendar_auto_on", {}, {"target": "admin"})
        await update.message.reply_text("ØªØ£ÙƒÙŠØ¯ ØªØ´ØºÙŠÙ„ Ø§Ù„ØªØ­Ø¯ÙŠØ« Ø§Ù„ØªÙ„Ù‚Ø§Ø¦ÙŠØŸ", reply_markup=kb_confirm())
        return

    if text == BTN_CAL_AUTO_OFF:
        if not is_super_admin(uid):
            await update.message.reply_text(deny_no_perm(), reply_markup=kb_admin(uid))
            return
        start_confirm(context, "calendar_auto_off", {}, {"target": "admin"})
        await update.message.reply_text("ØªØ£ÙƒÙŠØ¯ Ø¥ÙŠÙ‚Ø§Ù Ø§Ù„ØªØ­Ø¯ÙŠØ« Ø§Ù„ØªÙ„Ù‚Ø§Ø¦ÙŠØŸ", reply_markup=kb_confirm())
        return

    # ADMIN: groups menu
    if is_admin(uid) and text == BTN_GG_MENU:
        set_mode(context, MODE_GG_MENU)
        await update.message.reply_text("ğŸ‘¥ Ø¥Ø¯Ø§Ø±Ø© Ø§Ù„Ù‚Ø±ÙˆØ¨Ø§Øª:", reply_markup=kb_general_groups_admin())
        return

    if is_admin(uid) and text == BTN_GG_LIST:
        await update.message.reply_text(_gg_list_text(), reply_markup=kb_general_groups_admin())
        return

    if is_admin(uid) and text == BTN_GG_ADD:
        set_mode(context, MODE_GG_ADD)
        context.user_data[USER_STEP] = "name"
        context.user_data[USER_TEMP] = {}
        await update.message.reply_text("Ø§ÙƒØªØ¨ Ø§Ø³Ù… Ø§Ù„Ù‚Ø±ÙˆØ¨:", reply_markup=kb_wait_cancel())
        return

    if is_admin(uid) and text == BTN_GG_DEL:
        set_mode(context, MODE_GG_DEL)
        await update.message.reply_text("Ø£Ø±Ø³Ù„ ID Ø§Ù„Ù‚Ø±ÙˆØ¨ Ø§Ù„Ù…Ø±Ø§Ø¯ Ø­Ø°ÙÙ‡:", reply_markup=kb_wait_cancel())
        return

    if is_admin(uid) and text == BTN_GG_EDIT:
        set_mode(context, MODE_GG_EDIT_SELECT)
        await update.message.reply_text("Ø£Ø±Ø³Ù„ ID Ø§Ù„Ù‚Ø±ÙˆØ¨ Ø§Ù„Ù…Ø±Ø§Ø¯ ØªØ¹Ø¯ÙŠÙ„Ù‡:", reply_markup=kb_wait_cancel())
        return

    # ADMIN: contact service
    if is_admin(uid) and text == BTN_CC_MENU:
        set_mode(context, MODE_CC_MENU)
        await update.message.reply_text("ğŸ“² Ø®Ø¯Ù…Ø© Ø§Ù„Ø£Ø±Ù‚Ø§Ù…:", reply_markup=kb_contact_service_admin(uid))
        return

    if mode == MODE_CC_MENU and is_admin(uid):
        if text == BTN_CC_SHOW_OK:
            await update.message.reply_text(contacts_ok_text(), reply_markup=kb_contact_service_admin(uid))
            return
        if text == BTN_CC_SHOW_BAD:
            await update.message.reply_text(contacts_bad_text(), reply_markup=kb_contact_service_admin(uid))
            return

        if text == BTN_CC_EXPORT_XLSX:
            # âœ… ØªØµØ¯ÙŠØ± Excel Ù„Ù„Ø£Ø¯Ù…Ù†
            fn = f"contacts_export_{time.strftime('%Y%m%d_%H%M%S')}.xlsx"
            path = os.path.join(os.getcwd(), fn)
            try:
                await asyncio.to_thread(build_contacts_excel, path)
                await update.message.reply_document(document=open(path, "rb"), filename=fn, caption="ğŸ“¤ ØªÙ… ØªØµØ¯ÙŠØ± Ø§Ù„Ø£Ø±Ù‚Ø§Ù… Ø¨Ù†Ø¬Ø§Ø­.")
            except Exception as e:
                LOG.exception("Export excel failed: %s", e)
                await update.message.reply_text("âŒ ÙØ´Ù„ ØªØµØ¯ÙŠØ± Ø§Ù„Ù…Ù„Ù.", reply_markup=kb_contact_service_admin(uid))
            finally:
                try:
                    if os.path.exists(path):
                        os.remove(path)
                except Exception:
                    pass
            return

        if text == BTN_CC_ENABLE:
            if not is_super_admin(uid):
                await update.message.reply_text(deny_no_perm(), reply_markup=kb_contact_service_admin(uid))
                return
            start_confirm(context, "contact_toggle", {"enabled": True}, {"target": "admin"})
            await update.message.reply_text("ØªØ£ÙƒÙŠØ¯ ØªØ´ØºÙŠÙ„ Ø®Ø¯Ù…Ø© Ø§Ù„Ø£Ø±Ù‚Ø§Ù…ØŸ", reply_markup=kb_confirm())
            return

        if text == BTN_CC_DISABLE:
            if not is_super_admin(uid):
                await update.message.reply_text(deny_no_perm(), reply_markup=kb_contact_service_admin(uid))
                return
            start_confirm(context, "contact_toggle", {"enabled": False}, {"target": "admin"})
            await update.message.reply_text("ØªØ£ÙƒÙŠØ¯ Ø¥ÙŠÙ‚Ø§Ù Ø®Ø¯Ù…Ø© Ø§Ù„Ø£Ø±Ù‚Ø§Ù…ØŸ", reply_markup=kb_confirm())
            return

        if text == BTN_CC_CLEAR_OK:
            if not is_super_admin(uid):
                await update.message.reply_text(deny_no_perm(), reply_markup=kb_contact_service_admin(uid))
                return
            start_confirm(context, "contact_clear_ok", {}, {"target": "admin"})
            await update.message.reply_text("ØªØ£ÙƒÙŠØ¯ Ù…Ø³Ø­ Ø§Ù„Ø£Ø±Ù‚Ø§Ù… Ø§Ù„Ù…Ù‚Ø¨ÙˆÙ„Ø©ØŸ", reply_markup=kb_confirm())
            return

        if text == BTN_CC_CLEAR_BAD:
            if not is_super_admin(uid):
                await update.message.reply_text(deny_no_perm(), reply_markup=kb_contact_service_admin(uid))
                return
            start_confirm(context, "contact_clear_bad", {}, {"target": "admin"})
            await update.message.reply_text("ØªØ£ÙƒÙŠØ¯ Ù…Ø³Ø­ Ø§Ù„Ø£Ø±Ù‚Ø§Ù… Ø§Ù„Ù…Ø±ÙÙˆØ¶Ø©ØŸ", reply_markup=kb_confirm())
            return

    # SUPER: hide/rename/admin manage
    if text == BTN_HIDE_MENU:
        if not is_super_admin(uid):
            await update.message.reply_text(deny_no_perm(), reply_markup=kb_admin(uid))
            return
        set_mode(context, MODE_HIDE_MENU)
        await update.message.reply_text(hide_menu_text(), reply_markup=kb_wait_cancel())
        return

    if text == BTN_RENAME_MENU:
        if not is_super_admin(uid):
            await update.message.reply_text(deny_no_perm(), reply_markup=kb_admin(uid))
            return
        set_mode(context, MODE_RENAME_MENU)
        await update.message.reply_text(rename_menu_text(), reply_markup=kb_wait_cancel())
        return

    if text == BTN_AM_MENU:
        if not is_super_admin(uid):
            await update.message.reply_text(deny_no_perm(), reply_markup=kb_admin(uid))
            return
        set_mode(context, MODE_AM_MENU)
        await update.message.reply_text("ğŸ‘® Ø¥Ø¯Ø§Ø±Ø© Ø§Ù„Ø£Ø¯Ù…Ù†:", reply_markup=kb_admin_manage())
        return

    if mode == MODE_AM_MENU and is_super_admin(uid):
        if text == BTN_AM_LIST:
            admins = sorted(data_read(lambda: list(DATA.extra_admins)))
            msg = "ğŸ“‹ Ù„Ø§ ÙŠÙˆØ¬Ø¯ Ø£Ø¯Ù…Ù† Ù…Ø¶Ø§Ù." if not admins else ("ğŸ“‹ Ø§Ù„Ø£Ø¯Ù…Ù†:\n\n" + "\n".join([f"- {x}" for x in admins]))
            await update.message.reply_text(clip(msg), reply_markup=kb_admin_manage())
            return
        if text == BTN_AM_ADD:
            set_mode(context, MODE_AM_ADD)
            await update.message.reply_text("Ø£Ø±Ø³Ù„ ID Ø§Ù„Ø£Ø¯Ù…Ù† Ù„Ø¥Ø¶Ø§ÙØªÙ‡:", reply_markup=kb_wait_cancel())
            return
        if text == BTN_AM_DEL:
            set_mode(context, MODE_AM_DEL)
            await update.message.reply_text("Ø£Ø±Ø³Ù„ ID Ø§Ù„Ø£Ø¯Ù…Ù† Ù„Ø­Ø°ÙÙ‡:", reply_markup=kb_wait_cancel())
            return

    await update.message.reply_text("Ù…Ø§ ÙÙ‡Ù…Øª Ø·Ù„Ø¨Ùƒ ğŸ¤\nØ§Ø³ØªØ®Ø¯Ù… Ø§Ù„Ø£Ø²Ø±Ø§Ø±.", reply_markup=kb_main(uid))


# =========================================================
# ERROR HANDLER
# =========================================================
async def error_handler(update: object, context: ContextTypes.DEFAULT_TYPE) -> None:
    LOG.exception("Unhandled error: %s", context.error)


# =========================================================
# BUILD APP
# =========================================================
def build_app() -> Application:
    if not TOKEN or TOKEN == "8308362115:AAFj9WDYSjF0YYlvo1r1bgkRPyXi49h1VJ4":
        LOG.warning("âš ï¸ TOKEN Ù…Ø§Ø²Ø§Ù„ Ø§ÙØªØ±Ø§Ø¶ÙŠ. ØºÙŠÙ‘Ø±Ù‡ Ù‚Ø¨Ù„ Ø§Ù„ØªØ´ØºÙŠÙ„ Ø§Ù„Ø­Ù‚ÙŠÙ‚ÙŠ.")

    app = Application.builder().token(TOKEN).post_init(post_init).build()

    app.add_handler(CommandHandler("start", start_cmd))
    app.add_handler(CommandHandler("help", help_cmd))
    app.add_handler(CommandHandler("myid", myid_cmd))
    app.add_handler(CommandHandler("admin", admin_cmd))

    app.add_handler(MessageHandler(filters.CONTACT, handle_contact))
    app.add_handler(MessageHandler(filters.PHOTO | filters.Document.ALL, handle_media))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))

    app.add_error_handler(error_handler)
    setup_jobs(app)
    return app


def main() -> None:
    print("âœ… Ø§Ù„Ø¨ÙˆØª ÙŠØ¹Ù…Ù„ Ø§Ù„Ø¢Ù†... Ctrl+C Ù„Ù„Ø¥ÙŠÙ‚Ø§Ù.")
    app = build_app()
    app.run_polling()


if __name__ == "__main__":
    main()
